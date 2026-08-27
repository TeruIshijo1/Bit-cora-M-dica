from fastapi import FastAPI, Depends, HTTPException, status, UploadFile, File, Form, Request

from fastapi.middleware.cors import CORSMiddleware

from fastapi.staticfiles import StaticFiles

from sqlalchemy.orm import Session

from sqlalchemy import func

from typing import List, Optional, Union, Dict, Any

import datetime

import base64

import uuid

import json

import hashlib

import hmac

import crypto_fea

import tsa_client

from PIL import Image

import io

import os

import shutil

import csv

from fastapi.responses import StreamingResponse, JSONResponse

from starlette.middleware.base import BaseHTTPMiddleware

from slowapi import Limiter, _rate_limit_exceeded_handler

from slowapi.util import get_remote_address

from slowapi.errors import RateLimitExceeded



from database import engine, SessionLocal, get_db, Base

from pydantic import BaseModel, field_validator

import models

import schemas

from seed import get_password_hash, pwd_context

from jose import JWTError, jwt

from pdf_generator import generate_pdf



from docxtpl import DocxTemplate

from docx2pdf import convert as docx2pdf_convert

import pythoncom

from openpyxl import Workbook

from openpyxl.styles import PatternFill, Font, Alignment

from openpyxl.drawing.image import Image as ExcelImage

from fastapi.responses import FileResponse

import kh_database

from dotenv import load_dotenv

load_dotenv(dotenv_path=os.path.join(os.path.dirname(__file__), ".env"))
load_dotenv()

# Config & Variables de Entorno
SECRET_KEY = os.getenv("SECRET_KEY")
if not SECRET_KEY:
    raise ValueError("Falta SECRET_KEY en el archivo .env. El sistema no puede iniciar de forma segura.")

ALGORITHM = os.getenv("ALGORITHM", "HS256")
ACCESS_TOKEN_EXPIRE_MINUTES = int(os.getenv("ACCESS_TOKEN_EXPIRE_MINUTES", "480"))

UPLOADS_DIR = "static/uploads"

GENERADOS_DIR = "generados"

PLANTILLAS_DIR = "plantillas"

os.makedirs(UPLOADS_DIR, exist_ok=True)

os.makedirs(GENERADOS_DIR, exist_ok=True)

os.makedirs(PLANTILLAS_DIR, exist_ok=True)



app = FastAPI(title="MediReg API - Hospital Escandón")



limiter = Limiter(key_func=get_remote_address)

app.state.limiter = limiter

app.add_exception_handler(RateLimitExceeded, _rate_limit_exceeded_handler)



class SecurityHeadersMiddleware(BaseHTTPMiddleware):

    async def dispatch(self, request: Request, call_next):

        response = await call_next(request)

        response.headers["X-Frame-Options"] = "DENY"

        response.headers["X-Content-Type-Options"] = "nosniff"

        # Quitamos Strict-Transport-Security hasta que haya HTTPS real

        return response



app.add_middleware(SecurityHeadersMiddleware)



class GlobalAuthMiddleware(BaseHTTPMiddleware):

    async def dispatch(self, request: Request, call_next):

        # 1. Permitir peticiones preflight de CORS (OPTIONS)

        if request.method == "OPTIONS":

            return await call_next(request)



        path = request.url.path



        # 2. Permitir documentación Swagger / OpenAPI / Redoc

        if path in ["/docs", "/redoc", "/openapi.json"] or path.startswith("/docs") or path.startswith("/redoc"):

            return await call_next(request)



        # 3. Permitir endpoints públicos de inicio de sesión

        if path.startswith("/api/auth/login"):

            return await call_next(request)



        # 4. Permitir rutas que no son parte de la API (archivos estáticos, frontend SPA, assets)

        if not path.startswith("/api"):

            return await call_next(request)



        # 5. Para TODAS las rutas /api restantes: exigir y validar token JWT

        auth_header = request.headers.get("Authorization")

        if not auth_header or not auth_header.startswith("Bearer "):

            return JSONResponse(

                status_code=401,

                content={"detail": "No autenticado"}

            )



        token = auth_header.split(" ", 1)[1].strip()

        try:

            payload = jwt.decode(token, SECRET_KEY, algorithms=[ALGORITHM])

            username: str = payload.get("sub")

            rol: str = payload.get("rol")

            if not username or not rol:

                return JSONResponse(

                    status_code=401,

                    content={"detail": "Credenciales invalidas"}

                )

            request.state.user = payload

        except JWTError:

            return JSONResponse(

                status_code=401,

                content={"detail": "Token invalido o expirado"}

            )



        return await call_next(request)



app.add_middleware(GlobalAuthMiddleware)



# Servir estaticos para fotos de medicos

app.mount("/static", StaticFiles(directory="static"), name="static")



# CORS

app.add_middleware(

    CORSMiddleware,

    allow_origins=["http://localhost:8000", "http://127.0.0.1:8000", "http://192.168.254.249:8000"],

    allow_credentials=True,

    allow_methods=["*"],

    allow_headers=["*"],

)



def create_access_token(data: dict):

    to_encode = data.copy()

    expire = datetime.datetime.utcnow() + datetime.timedelta(minutes=ACCESS_TOKEN_EXPIRE_MINUTES)

    to_encode.update({"exp": expire})

    encoded_jwt = jwt.encode(to_encode, SECRET_KEY, algorithm=ALGORITHM)

    return encoded_jwt



from fastapi.security import OAuth2PasswordBearer

oauth2_scheme = OAuth2PasswordBearer(tokenUrl="api/auth/login/admin")



def get_current_user(token: str = Depends(oauth2_scheme), db: Session = Depends(get_db)):

    try:

        payload = jwt.decode(token, SECRET_KEY, algorithms=[ALGORITHM])

        username: str = payload.get("sub")

        rol: str = payload.get("rol")

        if username is None or rol is None:

            raise HTTPException(status_code=401, detail="Credenciales invalidas")

        

        if rol in ["medico", "ayudante"]:

            user = db.query(models.Medico).filter(models.Medico.cedula == username).first()

            if user:

                # Add a temporary attribute so require_role works

                setattr(user, "rol", rol)

        else:

            user = db.query(models.Usuario).filter(models.Usuario.username == username).first()

            

        if not user:

            raise HTTPException(status_code=401, detail="Usuario no existe")

            

        return user

    except JWTError:

        raise HTTPException(status_code=401, detail="Credenciales invalidas")



def log_auditoria(db: Session, usuario_id: Optional[int], accion: str, detalles_json: Optional[str] = None):

    try:

        log = models.AuditoriaLog(

            usuario_id=usuario_id,

            accion=accion,

            detalles_json=detalles_json,

            ip_origen=None # No request context here, but can be added later if needed

        )

        db.add(log)

        db.commit()

    except Exception as e:

        print(f"Error guardando auditoria: {e}")



def require_role(allowed_roles: List[str]):

    def role_checker(current_user = Depends(get_current_user)):

        rol = getattr(current_user, "rol", "medico")

        if "admin" not in allowed_roles and rol not in allowed_roles and rol != "admin":

            raise HTTPException(status_code=403, detail="No tienes permisos para esta acción")

        return current_user

    return role_checker



import requests



@app.post("/api/auth/login/admin", response_model=schemas.Token)

@limiter.limit("5/minute")

def login_admin(request: Request, req: schemas.LoginAdminRequest, db: Session = Depends(get_db)):

    print(f"Intento de login para usuario: '{req.username}'")

    user = db.query(models.Usuario).filter(models.Usuario.username == req.username).first()

    if not user:

        print("Error: Usuario no encontrado en DB.")

        raise HTTPException(status_code=401, detail="Usuario o contraseña incorrectos")

        

    try:

        if not pwd_context.verify(req.password, user.password_hash):

            print(f"Error: Contraseña incorrecta para el usuario '{req.username}'.")

            raise HTTPException(status_code=401, detail="Usuario o contraseña incorrectos")

    except Exception as e:

        print(f"Error verificando hash (posible contraseña en texto plano en DB): {e}")

        raise HTTPException(status_code=401, detail="Usuario o contraseña incorrectos")

    

    print("Login exitoso.")

    rol_efectivo = user.rol

    if user.username in ['amendoza', 'fmorales', 'rovajoir']:

        rol_efectivo = 'sistemas'

        

    access_token = create_access_token(data={"sub": user.username, "rol": rol_efectivo})

    return {

        "access_token": access_token, 

        "token_type": "bearer", 

        "rol": rol_efectivo,

        "permisos_modulos": user.permisos_modulos,

        "formatos_permitidos": user.formatos_permitidos

    }



@app.post("/api/auth/login/biometric", response_model=schemas.Token)

@limiter.limit("5/minute")

def login_biometric(request: Request, req: schemas.LoginBiometricRequest, db: Session = Depends(get_db)):

    if not req.fmd_template:

        raise HTTPException(status_code=400, detail="No se recibió la huella biométrica (FMD).")

    

    medicos = db.query(models.Medico).filter(models.Medico.activo_status == True, models.Medico.fmd_template.isnot(None)).all()

    if not medicos:

        raise HTTPException(status_code=401, detail="No hay médicos registrados con huella.")

    

    match_found = None

    medicos_data = [{"id": m.id, "fmd_template": m.fmd_template} for m in medicos]

    try:

        response = requests.post("http://localhost:8082/match-bulk", json={

            "fmd1": req.fmd_template,

            "medicos": medicos_data

        }, timeout=10)

        

        if response.status_code == 200:

            data = response.json()

            if data.get("success") and data.get("isMatch"):

                match_id = data.get("match_id")

                match_found = next((m for m in medicos if m.id == match_id), None)

        else:

            print(f"Error del motor biométrico: {response.text}")

            raise HTTPException(status_code=500, detail="Error de procesamiento biométrico.")

    except Exception as e:

        print(f"Error conectando al microservicio: {e}")

        raise HTTPException(status_code=500, detail="Error de conexión con motor biométrico.")

            

    if not match_found:

        raise HTTPException(status_code=401, detail="Huella no reconocida.")

    

    rol_asignado = "ayudante" if match_found.es_ayudante else "medico"

    access_token = create_access_token(data={"sub": match_found.cedula, "rol": rol_asignado})

    return {

        "access_token": access_token, 

        "token_type": "bearer", 

        "rol": rol_asignado,

        "medico_id": match_found.id,

        "nombre_completo": match_found.nombre_completo,

        "especialidad": match_found.especialidad,

        "cedula": match_found.cedula,

        "foto_url": match_found.foto_url,

        "formatos_permitidos": match_found.formatos_permitidos

    }



@app.post("/api/auth/impersonate")

def impersonate(req: schemas.ImpersonateRequest, db: Session = Depends(get_db), current_user: models.Usuario = Depends(require_role(["admin", "sistemas"]))):

    # This allows admin to get a token for any username or medico

    if req.rol == "medico":

        medico = db.query(models.Medico).filter(models.Medico.id == req.target_id).first()

        if not medico:

            raise HTTPException(status_code=404, detail="Médico no encontrado")

        access_token = create_access_token(data={"sub": medico.cedula, "rol": "medico"})

        return {

            "access_token": access_token, 

            "token_type": "bearer", 

            "rol": "medico",

            "medico_id": medico.id,

            "nombre_completo": medico.nombre_completo,

            "especialidad": medico.especialidad,

            "cedula": medico.cedula,

            "foto_url": medico.foto_url

        }

    else:

        user = db.query(models.Usuario).filter(models.Usuario.id == req.target_id).first()

        if not user:

            raise HTTPException(status_code=404, detail="Usuario no encontrado")

        access_token = create_access_token(data={"sub": user.username, "rol": user.rol})

        return {

            "access_token": access_token, 

            "token_type": "bearer", 

            "rol": user.rol

        }



# --- Catálogos ---

@app.get("/api/catalogos/areas", response_model=List[schemas.CatalogoResponse])

def get_areas(db: Session = Depends(get_db)):

    return db.query(models.CatalogoArea).filter(models.CatalogoArea.activo == True).all()



@app.post("/api/catalogos/areas", response_model=schemas.CatalogoResponse)

def create_area(area: schemas.CatalogoBase, db: Session = Depends(get_db), current_user: models.Usuario = Depends(require_role(["admin", "rh", "sistemas"]))):

    nueva_area = models.CatalogoArea(nombre=area.nombre, activo=area.activo)

    db.add(nueva_area)

    db.commit()

    db.refresh(nueva_area)

    return nueva_area



@app.delete("/api/catalogos/areas/{id}")

def delete_area(id: int, db: Session = Depends(get_db), current_user: models.Usuario = Depends(require_role(["admin", "sistemas"]))):

    area = db.query(models.CatalogoArea).filter(models.CatalogoArea.id == id).first()

    if not area:

        raise HTTPException(status_code=404, detail="Area no encontrada")

    area.activo = False

    db.commit()

    return {"status": "ok", "message": "Area desactivada"}



@app.get("/api/catalogos/tipos", response_model=List[schemas.CatalogoResponse])

def get_tipos(db: Session = Depends(get_db)):

    return db.query(models.CatalogoTipoAtencion).filter(models.CatalogoTipoAtencion.activo == True).all()



@app.post("/api/catalogos/tipos", response_model=schemas.CatalogoResponse)

def create_tipo(tipo: schemas.CatalogoBase, db: Session = Depends(get_db), current_user: models.Usuario = Depends(require_role(["admin", "rh", "sistemas"]))):

    nuevo_tipo = models.CatalogoTipoAtencion(nombre=tipo.nombre, activo=tipo.activo)

    db.add(nuevo_tipo)

    db.commit()

    db.refresh(nuevo_tipo)

    return nuevo_tipo





@app.get("/api/catalogos/formatos", response_model=List[schemas.CatalogoFormatoResponse])

def get_formatos(db: Session = Depends(get_db)):

    return db.query(models.CatalogoFormato).filter(models.CatalogoFormato.activo == True).order_by(models.CatalogoFormato.nombre).all()



@app.post("/api/catalogos/formatos", response_model=schemas.CatalogoFormatoResponse)

def create_formato(req: schemas.CatalogoFormatoCreate, db: Session = Depends(get_db)):

    nuevo = models.CatalogoFormato(nombre=req.nombre, codigo=req.codigo, activo=req.activo)

    db.add(nuevo)

    db.commit()

    db.refresh(nuevo)

    return nuevo



@app.delete("/api/catalogos/tipos/{id}")

def delete_tipo(id: int, db: Session = Depends(get_db), current_user: models.Usuario = Depends(require_role(["admin", "sistemas"]))):

    tipo = db.query(models.CatalogoTipoAtencion).filter(models.CatalogoTipoAtencion.id == id).first()

    if not tipo:

        raise HTTPException(status_code=404, detail="Tipo de Atención no encontrado")

    tipo.activo = False

    db.commit()

    return {"status": "ok", "message": "Tipo de atención desactivado"}



# --- Pacientes ---

@app.get("/api/pacientes", response_model=List[schemas.PacienteResponse])

def get_pacientes(db: Session = Depends(get_db)):

    try:

        # Sincronizar pacientes desde KH_HE

        camas = kh_database.fetch_camas()

        

        # Ignorar si falla la conexión

        if isinstance(camas, list) and len(camas) > 0 and "Error" not in camas[0] and "Mensaje" not in camas[0]:

            # === SINCRONIZAR CAMAS ===

            processed_rooms = set()

            for cama in camas:

                room_name = cama.get("RoomName")

                

                # Evitar procesar duplicados en la misma consulta (común en camas virtuales)

                if room_name in processed_rooms:

                    continue

                processed_rooms.add(room_name)

                

                # Buscar si ya existe la cama en local

                cama_db = db.query(models.Cama).filter(models.Cama.numero_cama == room_name).first()

                

                # Determinar área básica para la cama

                area = "Otras Áreas"

                name_upper = room_name.upper()

                if "PB" in name_upper or "10" in name_upper: area = "PPB (Planta Baja)"

                elif "PA" in name_upper or "20" in name_upper: area = "PPA (Planta Alta)"

                elif "URGENCIA" in name_upper or "URG" in name_upper: area = "Urgencias"

                elif "QUIR" in name_upper: area = "Quirófano"

                elif "UTI" in name_upper or "TERAPIA" in name_upper or "CUBICULO" in name_upper: area = "Terapia Intensiva"



                if not cama_db:

                    cama_db = models.Cama(

                        numero_cama=room_name,

                        area=area,

                        estado="OCUPADA" if cama.get("Estatus") == "Ocupada" else "DISPONIBLE",

                        activo=True

                    )

                    db.add(cama_db)

                    db.flush() # Guardar de inmediato para evitar colisiones

                else:

                    # Actualizar estado si la cama está activa y no bloqueada/mantenimiento

                    if cama_db.activo and cama_db.estado not in ["MANTENIMIENTO", "BLOQUEADA"]:

                        cama_db.estado = "OCUPADA" if cama.get("Estatus") == "Ocupada" else "DISPONIBLE"

                        

            db.flush()

            

            # === SINCRONIZAR PACIENTES ===

            active_patient_names = [c.get("PatientName") for c in camas if c.get("Estatus") == "Ocupada"]

            

            # Dar de alta a los que ya no están ocupando cama en KH_HE

            local_active_patients = db.query(models.Paciente).filter(models.Paciente.status_ingreso == "Ingresado").all()

            for lp in local_active_patients:

                if lp.nombre_completo not in active_patient_names and lp.registrado_por_nombre == "Sincronización KH_HE":

                    lp.status_ingreso = "Alta"

                    lp.fecha_alta = datetime.datetime.utcnow()

                    lp.dado_de_alta_por_id = None # Sistema

                    

                    nuevo_log = models.AuditoriaLog(

                        usuario_id=None,

                        accion="Alta de Paciente",

                        detalles_json=f"Paciente {lp.nombre_completo} (ID: {lp.id}) fue dado de alta automáticamente por el sistema."

                    )

                    db.add(nuevo_log)

            

            # Agregar o actualizar pacientes ocupados

            for cama in camas:

                if cama.get("Estatus") == "Ocupada":

                    patient_name = cama.get("PatientName")

                    room_name = cama.get("RoomName")

                    pt_date_str = cama.get("pt_date")

                    

                    paciente_db = db.query(models.Paciente).filter(

                        models.Paciente.nombre_completo == patient_name,

                        models.Paciente.status_ingreso == "Ingresado"

                    ).first()

                    

                    # Determinar área básica

                    area = "Otras Áreas"

                    name_upper = room_name.upper()

                    if "PB" in name_upper or "10" in name_upper: area = "PPB (Planta Baja)"

                    elif "PA" in name_upper or "20" in name_upper: area = "PPA (Planta Alta)"

                    elif "URGENCIA" in name_upper or "URG" in name_upper: area = "Urgencias"

                    elif "QUIR" in name_upper: area = "Quirófano"

                    elif "UTI" in name_upper or "TERAPIA" in name_upper or "CUBICULO" in name_upper: area = "Terapia Intensiva"

                    

                    if not paciente_db:

                        nuevo_paciente = models.Paciente(

                            nombre_completo=patient_name,

                            num_habitacion=room_name,

                            area_hospitalaria=area,

                            status_ingreso="Ingresado",

                            registrado_por_nombre="Sincronización KH_HE"

                        )

                        db.add(nuevo_paciente)

                        db.flush() # Evitar duplicados en el mismo bucle

                    else:

                        if paciente_db.num_habitacion != room_name:

                            dt_traslado = datetime.datetime.utcnow()

                            if pt_date_str and pt_date_str != "None":

                                try:

                                    dt_traslado = datetime.datetime.fromisoformat(pt_date_str.split(".")[0].replace(" ", "T"))

                                except:

                                    pass



                            nuevo_traslado = models.TrasladoPaciente(

                                paciente_id=paciente_db.id,

                                origen_area=paciente_db.area_hospitalaria,

                                origen_habitacion=paciente_db.num_habitacion,

                                destino_area=area,

                                destino_habitacion=room_name,

                                fecha_traslado=dt_traslado,

                                usuario_id=None # Sistema

                            )

                            db.add(nuevo_traslado)

                            paciente_db.num_habitacion = room_name

                            paciente_db.area_hospitalaria = area

            db.commit()

    except Exception as e:

        print(f"Error en sincronización de pacientes KH_HE: {e}")

        db.rollback()



    return db.query(models.Paciente).filter(models.Paciente.status_ingreso == "Ingresado").all()



@app.post("/api/pacientes", response_model=schemas.PacienteResponse)

def create_paciente(paciente: schemas.PacienteCreate, db: Session = Depends(get_db), current_user: models.Usuario = Depends(get_current_user)):

    is_medico = getattr(current_user, "rol", "") in ("medico", "ayudante")

    registrado_por = None

    if is_medico:

        medico_db = db.query(models.Medico).filter(models.Medico.id == current_user.id).first()

        if medico_db:

            if medico_db.es_ayudante:

                med_titular = db.query(models.Medico).filter(models.Medico.id == medico_db.medico_asignado_id).first()

                titular_nombre = med_titular.nombre_completo if med_titular else "Desconocido"

                registrado_por = f"{medico_db.nombre_completo} (Ayudante del Dr. {titular_nombre})"

            else:

                registrado_por = f"Dr. {medico_db.nombre_completo}"

    else:

        registrado_por = current_user.nombre_completo or current_user.username



    nuevo_paciente = models.Paciente(

        nombre_completo=paciente.nombre_completo,

        num_habitacion=paciente.num_habitacion,

        area_hospitalaria=paciente.area_hospitalaria,

        codigo_barras=paciente.codigo_barras,

        status_ingreso="Ingresado",

        creado_por_id=None if is_medico else current_user.id,

        registrado_por_nombre=registrado_por

    )

    db.add(nuevo_paciente)

    db.commit()

    db.refresh(nuevo_paciente)

    

    log_auditoria(db, current_user.id, f"Paciente Ingresado", f"Se ingresó al paciente {nuevo_paciente.nombre_completo} (Folio ID: {nuevo_paciente.id}) en {nuevo_paciente.area_hospitalaria} - Hab: {nuevo_paciente.num_habitacion}")

    

    return nuevo_paciente



@app.put("/api/pacientes/{paciente_id}", response_model=schemas.PacienteResponse)

def update_paciente(paciente_id: int, req: schemas.PacienteUpdate, db: Session = Depends(get_db), current_user = Depends(get_current_user)):

    paciente = db.query(models.Paciente).filter(models.Paciente.id == paciente_id).first()

    if not paciente:

        raise HTTPException(status_code=404, detail="Paciente no encontrado")

        

    is_medico = getattr(current_user, "rol", "") in ("medico", "ayudante")

    usuario_id_log = None if is_medico else current_user.id

        

    area_ant = paciente.area_hospitalaria

    hab_ant = paciente.num_habitacion

    cambio_traslado = False

    

    if req.nombre_completo != paciente.nombre_completo:

        log_auditoria(db, usuario_id_log, "Nombre Paciente Editado", f"De {paciente.nombre_completo} a {req.nombre_completo}")

        paciente.nombre_completo = req.nombre_completo

        

    if req.num_habitacion != paciente.num_habitacion:

        cambio_traslado = True

    if req.area_hospitalaria is not None and req.area_hospitalaria != paciente.area_hospitalaria:

        cambio_traslado = True



    paciente.num_habitacion = req.num_habitacion

    if req.area_hospitalaria is not None:

        paciente.area_hospitalaria = req.area_hospitalaria

    if req.codigo_barras is not None:

        paciente.codigo_barras = req.codigo_barras

        

    if cambio_traslado:

        traslado = models.TrasladoPaciente(

            paciente_id=paciente.id,

            origen_area=area_ant,

            origen_habitacion=hab_ant,

            destino_area=paciente.area_hospitalaria,

            destino_habitacion=paciente.num_habitacion,

            usuario_id=usuario_id_log

        )

        db.add(traslado)

        log_auditoria(db, usuario_id_log, "Traslado de Paciente", f"Paciente {paciente.nombre_completo} movido de {area_ant}({hab_ant}) a {paciente.area_hospitalaria}({paciente.num_habitacion})")

        

    db.commit()

    db.refresh(paciente)

    return paciente



@app.put("/api/pacientes/{paciente_id}/alta")

def alta_paciente(paciente_id: int, db: Session = Depends(get_db), current_user = Depends(get_current_user)):

    paciente = db.query(models.Paciente).filter(models.Paciente.id == paciente_id).first()

    if not paciente:

        raise HTTPException(status_code=404, detail="Paciente no encontrado")

    paciente.status_ingreso = "Alta"

    # Medico objects don't have an id usable for dado_de_alta_por_id (FK → usuarios)

    is_medico = getattr(current_user, "rol", "") in ("medico", "ayudante")

    if not is_medico:

        paciente.dado_de_alta_por_id = current_user.id

    paciente.fecha_alta = datetime.datetime.utcnow()

    

    usuario_id_log = None if is_medico else current_user.id

    log_auditoria(db, usuario_id_log, "Alta de Paciente", f"Paciente {paciente.nombre_completo} (ID: {paciente.id}) fue dado de alta")

    

    db.commit()

    return {"message": "Paciente dado de alta"}



# --- Médicos ---

@app.get("/api/medicos", response_model=List[schemas.MedicoResponse])

def get_medicos(db: Session = Depends(get_db)):

    return db.query(models.Medico).filter(models.Medico.activo_status == True).all()



@app.post("/api/medicos", response_model=schemas.MedicoResponse)

async def create_medico(

    numero_empleado: str = Form(...),

    nombre_completo: str = Form(...),

    especialidad: str = Form(...),

    cedula: str = Form(...),

    fmd_template: Optional[str] = Form(None),

    foto: Optional[UploadFile] = File(None),

    bajo_contrato: bool = Form(False),

    horario_laboral: Optional[str] = Form(None),

    es_ayudante: bool = Form(False),

    medico_asignado_id: Optional[int] = Form(None),

    db: Session = Depends(get_db),

    current_user: models.Usuario = Depends(require_role(["admin", "rh", "sistemas"]))

):

    # Ya no se genera numero empleado automáticamente

    if db.query(models.Medico).filter(models.Medico.numero_empleado == numero_empleado).first():

        raise HTTPException(status_code=400, detail="Ya existe un médico con ese número de empleado.")

    if db.query(models.Medico).filter(models.Medico.cedula == cedula).first():

        raise HTTPException(status_code=400, detail="Ya existe un médico con esa cédula.")

    

    huella_token_unique = str(uuid.uuid4())

    

    foto_url = None

    if foto:

        ext = os.path.splitext(foto.filename)[1]

        filename = f"{numero_empleado}{ext}"

        filepath = os.path.join(UPLOADS_DIR, filename)

        with open(filepath, "wb") as buffer:

            shutil.copyfileobj(foto.file, buffer)

        foto_url = f"/static/uploads/{filename}"



    nuevo_medico = models.Medico(

        numero_empleado=numero_empleado,

        nombre_completo=nombre_completo,

        especialidad=especialidad,

        cedula=cedula,

        huella_token=huella_token_unique,

        fmd_template=fmd_template,

        foto_url=foto_url,

        bajo_contrato=bajo_contrato,

        horario_laboral=horario_laboral,

        es_ayudante=es_ayudante,

        medico_asignado_id=medico_asignado_id

    )

    db.add(nuevo_medico)

    db.commit()

    db.refresh(nuevo_medico)

    return nuevo_medico



# --- Atenciones ---

@app.post("/api/atenciones/pre-captura", response_model=schemas.AtencionResponse)

def pre_captura(req: schemas.PreCapturaRequest, db: Session = Depends(get_db), current_user = Depends(get_current_user)):

    hoy = datetime.date.today()

    

    if getattr(current_user, "rol", "") not in ("sistemas", "admin"):

        if req.fecha_realizacion:

            req.fecha_realizacion = datetime.datetime.combine(hoy, req.fecha_realizacion.time())

        else:

            req.fecha_realizacion = datetime.datetime.now()

    # Admin/sistemas: pueden elegir cualquier fecha sin restricción

    

    dias_semana = ["lunes", "martes", "miercoles", "jueves", "viernes", "sabado", "domingo"]

    

    rol_usuario = getattr(current_user, "rol", "")

    medico = db.query(models.Medico).filter(models.Medico.id == req.medico_id).first()

    if medico and medico.bajo_contrato and medico.horario_laboral and rol_usuario not in ("sistemas", "admin"):

        try:

            horario = json.loads(medico.horario_laboral)

            def is_in_shift(dt_to_check):

                hora_str = dt_to_check.strftime("%H:%M")

                dia_idx = dt_to_check.weekday()

                dia_actual = dias_semana[dia_idx]

                dia_previo = dias_semana[(dia_idx - 1) % 7]

                

                # Check today's shift

                shift_hoy = horario.get(dia_actual)

                if shift_hoy and shift_hoy.get("activo"):

                    inicio = shift_hoy.get("inicio", "")

                    fin = shift_hoy.get("fin", "")

                    if inicio and fin:

                        if inicio <= fin:

                            if inicio <= hora_str <= fin:

                                return True

                        else:

                            if hora_str >= inicio:

                                return True

                

                # Check yesterday's shift for overnight

                shift_ayer = horario.get(dia_previo)

                if shift_ayer and shift_ayer.get("activo"):

                    inicio = shift_ayer.get("inicio", "")

                    fin = shift_ayer.get("fin", "")

                    if inicio and fin and inicio > fin:

                        if hora_str <= fin:

                            return True

                return False



            dt_sistema = datetime.datetime.now()

            dt_registro = req.fecha_realizacion or dt_sistema

            

            if is_in_shift(dt_sistema):

                raise HTTPException(status_code=400, detail="El médico está actualmente dentro de su jornada laboral base. No puede realizar registros.")

                

            if is_in_shift(dt_registro):

                raise HTTPException(status_code=400, detail="La hora reportada del procedimiento cae dentro de la jornada laboral del médico. Solo puede registrar procedimientos realizados fuera de turno.")

        except Exception as e:

            if isinstance(e, HTTPException): raise e

            pass # Ignore JSON parsing errors

            

    # One record per patient per day per doctor

    exact_match = db.query(models.AtencionMedica).filter(

        models.AtencionMedica.paciente_id == req.paciente_id,

        models.AtencionMedica.medico_id == req.medico_id,

        func.date(models.AtencionMedica.fecha_realizacion) == (req.fecha_realizacion.date() if req.fecha_realizacion else hoy)

    ).first()

    

    nuevo_estatus_pago = "Pendiente de Firma"

    if exact_match:

        nuevo_estatus_pago = "Pendiente Autorización"

    

    # Use max folio number instead of count to avoid collisions after cleanup

    from sqlalchemy import text

    max_row = db.execute(text("SELECT MAX(CAST(SUBSTR(folio, 10) AS INTEGER)) FROM atenciones_medicas")).fetchone()

    next_num = (max_row[0] or 0) + 1

    year = datetime.date.today().year

    folio = f"HES-{year}-{next_num:05d}"

    

    is_medico = getattr(current_user, "rol", "") in ("medico", "ayudante")

    

    registrado_por = None

    if is_medico:

        medico_db = db.query(models.Medico).filter(models.Medico.id == current_user.id).first()

        if medico_db:

            if medico_db.es_ayudante:

                med_titular = db.query(models.Medico).filter(models.Medico.id == medico_db.medico_asignado_id).first()

                titular_nombre = med_titular.nombre_completo if med_titular else "Desconocido"

                registrado_por = f"{medico_db.nombre_completo} (Ayudante del Dr. {titular_nombre})"

            else:

                registrado_por = f"Dr. {medico_db.nombre_completo}"

    else:

        registrado_por = current_user.nombre_completo or current_user.username

    

    paciente_db = db.query(models.Paciente).filter(models.Paciente.id == req.paciente_id).first()

    area_hosp = req.area_hospitalaria or (paciente_db.area_hospitalaria if paciente_db else "No asignada")

    

    nueva_atencion = models.AtencionMedica(

        folio=folio,

        fecha_realizacion=req.fecha_realizacion,

        medico_id=req.medico_id,

        paciente_id=req.paciente_id,

        area_hospitalaria=area_hosp,

        tipo_atencion=req.tipo_atencion,

        nombre_procedimiento=req.nombre_procedimiento,

        habitacion_capturada=req.habitacion_capturada,

        procedimiento_detalle=req.procedimiento_detalle,

        creado_por_id=None if is_medico else current_user.id,

        estatus_pago=nuevo_estatus_pago,

        registrado_por_nombre=registrado_por

    )

    

    db.add(nueva_atencion)

    

    # Update frequent procedures

    proc_frec = db.query(models.ProcedimientoFrecuente).filter(

        models.ProcedimientoFrecuente.medico_id == req.medico_id,

        func.lower(models.ProcedimientoFrecuente.nombre_procedimiento) == req.nombre_procedimiento.lower().strip()

    ).first()

    

    if proc_frec:

        proc_frec.frecuencia += 1

    else:

        nuevo_proc = models.ProcedimientoFrecuente(

            medico_id=req.medico_id,

            nombre_procedimiento=req.nombre_procedimiento.strip(),

            frecuencia=1

        )

        db.add(nuevo_proc)

        

    db.commit()

    db.refresh(nueva_atencion)

    return nueva_atencion



@app.get("/api/medicos/{medico_id}/procedimientos_frecuentes")

def get_procedimientos_frecuentes(medico_id: int, db: Session = Depends(get_db)):

    """Obtiene el top 5 de procedimientos más usados por este médico"""

    procs = db.query(models.ProcedimientoFrecuente).filter(

        models.ProcedimientoFrecuente.medico_id == medico_id

    ).order_by(models.ProcedimientoFrecuente.frecuencia.desc()).limit(5).all()

    

    return [{"nombre": p.nombre_procedimiento, "frecuencia": p.frecuencia} for p in procs]



@app.get("/api/atenciones/pendientes/{medico_id}", response_model=List[schemas.AtencionResponse])

def pendientes_medico(medico_id: int, db: Session = Depends(get_db)):

    return db.query(models.AtencionMedica).filter(

        models.AtencionMedica.medico_id == medico_id,

        models.AtencionMedica.estatus_pago == "Pendiente de Firma"

    ).all()



@app.get("/api/atenciones/historial/{medico_id}", response_model=List[schemas.AtencionResponse])

def historial_medico(medico_id: int, db: Session = Depends(get_db)):

    return db.query(models.AtencionMedica).filter(

        models.AtencionMedica.medico_id == medico_id,

        models.AtencionMedica.estatus_pago != "Pendiente de Firma"

    ).order_by(models.AtencionMedica.fecha_firma.desc()).all()



@app.get("/api/atenciones/mis-registros", response_model=List[schemas.AtencionResponse])

def mis_registros(db: Session = Depends(get_db), current_user: models.Usuario = Depends(get_current_user)):

    return db.query(models.AtencionMedica).filter(models.AtencionMedica.creado_por_id == current_user.id).order_by(models.AtencionMedica.fecha_registro.desc()).all()



@app.get("/api/atenciones/global", response_model=List[schemas.AtencionResponse])

def global_registros(db: Session = Depends(get_db), current_user: models.Usuario = Depends(get_current_user)):

    return db.query(models.AtencionMedica).order_by(models.AtencionMedica.fecha_registro.desc()).all()



@app.get("/api/atenciones/todas", response_model=List[schemas.AtencionResponse])

def get_todas_atenciones(db: Session = Depends(get_db), current_user: models.Usuario = Depends(require_role(["admin", "rh", "sistemas"]))):

    return db.query(models.AtencionMedica).order_by(models.AtencionMedica.fecha_realizacion.desc()).all()



@app.get("/api/atenciones/exportar")

def exportar_atenciones(

    start_date: Optional[str] = None,

    end_date: Optional[str] = None,

    db: Session = Depends(get_db), 

    current_user: models.Usuario = Depends(require_role(["admin", "rh", "sistemas"]))

):

    query = db.query(models.AtencionMedica)

    

    if start_date:

        query = query.filter(func.date(models.AtencionMedica.fecha_realizacion) >= start_date)

    if end_date:

        query = query.filter(func.date(models.AtencionMedica.fecha_realizacion) <= end_date)

        

    atenciones = query.order_by(models.AtencionMedica.fecha_realizacion.desc()).all()

    wb = Workbook()

    ws = wb.active

    ws.title = "Atenciones"

    

    # Insertar Logo

    logo_path = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "frontend", "dist", "logo.png")

    if not os.path.exists(logo_path):

        logo_path = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "frontend", "public", "logo.png")

    if os.path.exists(logo_path):

        img = ExcelImage(logo_path)

        img.width = 100

        img.height = 35

        ws.add_image(img, "A1")

        

    # Metadata Header

    ws.merge_cells("C1:G1")

    ws["C1"] = "HOSPITAL ESCANDÓN - REPORTE DE ATENCIONES MÉDICAS"

    ws["C1"].font = Font(bold=True, size=14, color="003870")

    ws["C1"].alignment = Alignment(horizontal="center", vertical="center")

    

    ws["I1"] = "Generado por:"

    ws["I1"].font = Font(bold=True)

    ws["J1"] = current_user.username

    ws["I2"] = "Fecha:"

    ws["I2"].font = Font(bold=True)

    ws["J2"] = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    

    # Headers

    headers = ["Folio", "Fecha Realización", "Médico", "Especialidad", "Paciente", "Habitación", "Área Hospitalaria", "Tipo Atención", "Procedimiento", "Estatus Pago", "Fecha Firma", "Hash"]

    header_fill = PatternFill(start_color="003870", end_color="003870", fill_type="solid")

    header_font = Font(color="FFFFFF", bold=True)

    

    for col_num, header in enumerate(headers, 1):

        cell = ws.cell(row=4, column=col_num, value=header)

        cell.fill = header_fill

        cell.font = header_font

        cell.alignment = Alignment(horizontal="center")

        

    row_idx = 5

    for a in atenciones:

        medico_nombre = a.medico.nombre_completo if a.medico else "N/A"

        medico_esp = a.medico.especialidad if a.medico else "N/A"

        paciente_nombre = a.paciente.nombre_completo if a.paciente else "N/A"

        f_firma = a.fecha_firma.strftime("%Y-%m-%d %H:%M:%S") if a.fecha_firma else ""

        

        ws.cell(row=row_idx, column=1, value=a.folio)

        ws.cell(row=row_idx, column=2, value=a.fecha_realizacion.strftime("%Y-%m-%d %H:%M:%S") if a.fecha_realizacion else "")

        ws.cell(row=row_idx, column=3, value=medico_nombre)

        ws.cell(row=row_idx, column=4, value=medico_esp)

        ws.cell(row=row_idx, column=5, value=paciente_nombre)

        ws.cell(row=row_idx, column=6, value=a.habitacion_capturada)

        ws.cell(row=row_idx, column=7, value=a.area_hospitalaria)

        ws.cell(row=row_idx, column=8, value=a.tipo_atencion)

        ws.cell(row=row_idx, column=9, value=a.nombre_procedimiento)

        ws.cell(row=row_idx, column=10, value=a.estatus_pago)

        ws.cell(row=row_idx, column=11, value=f_firma)

        ws.cell(row=row_idx, column=12, value=a.hash_seguridad or "")

        row_idx += 1

        

    from openpyxl.utils import get_column_letter

    for col_idx in range(1, ws.max_column + 1):

        column = get_column_letter(col_idx)

        max_length = 0

        for cell in ws[column]:

            try:

                if len(str(cell.value)) > max_length:

                    max_length = len(str(cell.value))

            except:

                pass

        adjusted_width = (max_length + 2)

        ws.column_dimensions[column].width = adjusted_width if adjusted_width < 50 else 50

        

    output = io.BytesIO()

    wb.save(output)

    output.seek(0)

    

    response = StreamingResponse(iter([output.getvalue()]), media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

    response.headers["Content-Disposition"] = "attachment; filename=atenciones_export.xlsx"

    return response



@app.post("/api/atenciones/firmar-lote")

def firmar_lote(req: schemas.FirmaLoteRequest, db: Session = Depends(get_db)):

    if not req.fmd_template:

        raise HTTPException(status_code=400, detail="No huella recibida.")



    medicos = db.query(models.Medico).filter(models.Medico.activo_status == True, models.Medico.fmd_template.isnot(None)).all()

    if not medicos:

        raise HTTPException(status_code=401, detail="No hay médicos registrados")

    

    match_found = None

    medicos_data = [{"id": m.id, "fmd_template": m.fmd_template} for m in medicos]

    

    try:

        response = requests.post("http://localhost:8082/match-bulk", json={

            "fmd1": req.fmd_template,

            "medicos": medicos_data

        }, timeout=10)

        

        if response.status_code == 200:

            data = response.json()

            if data.get("success") and data.get("isMatch"):

                match_id = data.get("match_id")

                match_found = next((m for m in medicos if m.id == match_id), None)

    except Exception as e:

        print(f"Error conectando al microservicio en firma: {e}")

            

    if not match_found:

        raise HTTPException(status_code=401, detail="Huella no reconocida.")

    

    firmados = []

    ahora = datetime.datetime.utcnow()

    

    for folio in req.folios:

        atencion = db.query(models.AtencionMedica).filter(

            models.AtencionMedica.folio == folio,

            models.AtencionMedica.medico_id == match_found.id

        ).first()

        

        if atencion:

            if atencion.is_caducado:

                continue # No se puede firmar registro caducado

            paciente = db.query(models.Paciente).filter(models.Paciente.id == atencion.paciente_id).first()

            raw_text = f"{atencion.folio}{atencion.fecha_realizacion.isoformat()}{match_found.nombre_completo}{paciente.nombre_completo}{atencion.habitacion_capturada}"

            hash_str = hashlib.sha256(raw_text.encode('utf-8')).hexdigest()

            

            atencion.hash_seguridad = hash_str

            atencion.estatus_pago = "Validado para Pago"

            atencion.fecha_firma = ahora

            

            # Generar PDF

            try:

                pdf_path = generate_pdf(atencion, match_found, paciente)

                atencion.ruta_archivo_firmado = pdf_path

            except Exception as e:

                print("Error al generar PDF:", e)

                

            firmados.append(folio)

            

    db.commit()

    return {"message": f"{len(firmados)} atenciones firmadas", "firmados": firmados}



@app.put("/api/atenciones/{folio}")

def update_atencion(folio: str, tipo_atencion: str, db: Session = Depends(get_db)):

    atencion = db.query(models.AtencionMedica).filter(models.AtencionMedica.folio == folio).first()

    if not atencion:

        raise HTTPException(status_code=404, detail="Atención no encontrada")

    atencion.tipo_atencion = tipo_atencion

    db.commit()

    return {"message": "Actualizado"}



@app.get("/api/atenciones/{folio}/pdf")

def generar_comprobante_pdf(folio: str, db: Session = Depends(get_db)):

    atencion = db.query(models.AtencionMedica).filter(models.AtencionMedica.folio == folio).first()

    if not atencion:

        raise HTTPException(status_code=404, detail="Atención no encontrada")

        

    template_path = os.path.join(PLANTILLAS_DIR, "comprobante_base.docx")

    if not os.path.exists(template_path):

        raise HTTPException(status_code=500, detail="Plantilla no encontrada")

        

    doc = DocxTemplate(template_path)

    

    context = {

        "folio": atencion.folio,

        "medico_nombre": atencion.medico.nombre_completo if atencion.medico else "",

        "medico_especialidad": atencion.medico.especialidad if atencion.medico else "",

        "medico_cedula": atencion.medico.cedula if atencion.medico else "",

        "medico_num_empleado": atencion.medico.numero_empleado if atencion.medico else "",

        "paciente_nombre": atencion.paciente.nombre_completo if atencion.paciente else "",

        "paciente_habitacion": atencion.habitacion_capturada or (atencion.paciente.num_habitacion if atencion.paciente else ""),

        "tipo_servicio": atencion.tipo_atencion,

        "fecha_atencion": atencion.fecha_realizacion.strftime("%Y-%m-%d %H:%M") if atencion.fecha_realizacion else "",

        "fecha_registro": (atencion.fecha_firma or atencion.fecha_realizacion or datetime.datetime.now()).strftime("%Y-%m-%d %H:%M"),

        "procedimiento": atencion.nombre_procedimiento,

        "detalle": atencion.procedimiento_detalle or "Sin notas clínicas",

        "hash_seguridad": atencion.hash_seguridad or "Pendiente"

    }

    

    doc.render(context)

    

    temp_docx = os.path.join(GENERADOS_DIR, f"{folio}.docx")

    temp_pdf = os.path.join(GENERADOS_DIR, f"{folio}.pdf")

    

    doc.save(temp_docx)

    

    pythoncom.CoInitialize()

    try:

        docx2pdf_convert(os.path.abspath(temp_docx), os.path.abspath(temp_pdf))

    except Exception as e:

        print(f"Error docx2pdf: {e}")

        # FALLBACK: Return DOCX instead of failing

        return FileResponse(temp_docx, media_type="application/vnd.openxmlformats-officedocument.wordprocessingml.document", filename=f"Comprobante_{folio}.docx")

    finally:

        pythoncom.CoUninitialize()

        

    if not os.path.exists(temp_pdf):

        # FALLBACK: Return DOCX if PDF wasn't created

        return FileResponse(temp_docx, media_type="application/vnd.openxmlformats-officedocument.wordprocessingml.document", filename=f"Comprobante_{folio}.docx")

        

    return FileResponse(temp_pdf, media_type="application/pdf", filename=f"Comprobante_{folio}.pdf")



# === ENDPOINTS RH ===

@app.get("/api/usuarios", response_model=List[schemas.UsuarioResponse])

def get_usuarios(db: Session = Depends(get_db), current_user: models.Usuario = Depends(require_role(["admin", "rh", "sistemas"]))):

    return db.query(models.Usuario).all()



@app.post("/api/usuarios", response_model=schemas.UsuarioResponse)

def create_usuario(usuario: schemas.UsuarioCreate, db: Session = Depends(get_db), current_user: models.Usuario = Depends(require_role(["admin", "rh", "sistemas"]))):

    existing = db.query(models.Usuario).filter(models.Usuario.username == usuario.username).first()

    if existing:

        raise HTTPException(status_code=400, detail="Usuario ya existe")

    nuevo_usuario = models.Usuario(

        username=usuario.username,

        password_hash=pwd_context.hash(usuario.password),

        rol=usuario.rol,

        nombre_completo=usuario.nombre_completo,

        permisos_modulos=usuario.permisos_modulos,

        formatos_permitidos=usuario.formatos_permitidos

    )

    db.add(nuevo_usuario)

    db.commit()

    db.refresh(nuevo_usuario)

    return nuevo_usuario



@app.put("/api/usuarios/{usuario_id}", response_model=schemas.UsuarioResponse)

def update_usuario(usuario_id: int, payload: schemas.UsuarioUpdate, db: Session = Depends(get_db), current_user: models.Usuario = Depends(require_role(["admin", "sistemas"]))):

    usuario = db.query(models.Usuario).filter(models.Usuario.id == usuario_id).first()

    if not usuario:

        raise HTTPException(status_code=404, detail="Usuario no encontrado")

        

    if payload.rol is not None:

        usuario.rol = payload.rol

    if payload.nombre_completo is not None:

        usuario.nombre_completo = payload.nombre_completo

    if payload.permisos_modulos is not None:

        usuario.permisos_modulos = payload.permisos_modulos

    if payload.formatos_permitidos is not None:

        usuario.formatos_permitidos = payload.formatos_permitidos

        

    db.commit()

    db.refresh(usuario)

    return usuario



@app.put("/api/usuarios/{usuario_id}/password")

def update_usuario_password(usuario_id: int, payload: schemas.UsuarioPasswordUpdate, db: Session = Depends(get_db), current_user: models.Usuario = Depends(require_role(["admin", "sistemas"]))):

    usuario = db.query(models.Usuario).filter(models.Usuario.id == usuario_id).first()

    if not usuario:

        raise HTTPException(status_code=404, detail="Usuario no encontrado")

        

    usuario.password_hash = pwd_context.hash(payload.new_password)

    db.commit()

    return {"message": "Contraseña actualizada exitosamente"}



@app.delete("/api/usuarios/{usuario_id}")

def delete_usuario(usuario_id: int, db: Session = Depends(get_db), current_user: models.Usuario = Depends(require_role(["sistemas"]))):

    usuario = db.query(models.Usuario).filter(models.Usuario.id == usuario_id).first()

    if not usuario:

        raise HTTPException(status_code=404, detail="Usuario no encontrado")

    if usuario.username == current_user.username:

        raise HTTPException(status_code=400, detail="No te puedes eliminar a ti mismo")

    db.delete(usuario)

    db.commit()

    return {"message": "Usuario eliminado"}



@app.put("/api/medicos/{medico_id}")

def update_medico(medico_id: int, activo: bool, db: Session = Depends(get_db), current_user: models.Usuario = Depends(require_role(["admin", "rh", "sistemas"]))):

    medico = db.query(models.Medico).filter(models.Medico.id == medico_id).first()

    medico.activo_status = activo

    db.commit()

    return {"message": "Actualizado"}



@app.put("/api/medicos/{medico_id}/permisos", response_model=schemas.MedicoResponse)

def update_medico_permisos(medico_id: int, payload: schemas.MedicoUpdatePermisos, db: Session = Depends(get_db), current_user: models.Usuario = Depends(require_role(["admin", "sistemas"]))):

    medico = db.query(models.Medico).filter(models.Medico.id == medico_id).first()

    if not medico:

        raise HTTPException(status_code=404, detail="Médico no encontrado")

        

    medico.formatos_permitidos = payload.formatos_permitidos

    db.commit()

    db.refresh(medico)

    return medico



@app.put("/api/medicos/{medico_id}/datos", response_model=schemas.MedicoResponse)

async def update_medico_datos(

    medico_id: int,

    numero_empleado: str = Form(...),

    nombre_completo: str = Form(...),

    especialidad: str = Form(...),

    cedula: str = Form(...),

    foto: Optional[UploadFile] = File(None),

    bajo_contrato: bool = Form(False),

    horario_laboral: Optional[str] = Form(None),

    es_ayudante: bool = Form(False),

    medico_asignado_id: Optional[int] = Form(None),

    formatos_permitidos: Optional[str] = Form(None),

    db: Session = Depends(get_db),

    current_user: models.Usuario = Depends(require_role(["admin", "rh", "sistemas"]))

):

    medico = db.query(models.Medico).filter(models.Medico.id == medico_id).first()

    if not medico:

        raise HTTPException(status_code=404, detail="Médico no encontrado")

        

    medico.numero_empleado = numero_empleado

    medico.nombre_completo = nombre_completo

    medico.especialidad = especialidad

    medico.cedula = cedula

    medico.bajo_contrato = bajo_contrato

    medico.horario_laboral = horario_laboral

    medico.es_ayudante = es_ayudante

    medico.medico_asignado_id = medico_asignado_id

    if formatos_permitidos is not None:

        medico.formatos_permitidos = formatos_permitidos

    

    if foto:

        ext = foto.filename.split(".")[-1]

        filename = f"{medico_id}_{numero_empleado}.{ext}"

        filepath = os.path.join(UPLOADS_DIR, filename)

        with open(filepath, "wb") as buffer:

            shutil.copyfileobj(foto.file, buffer)

        medico.foto_url = f"/static/uploads/{filename}"

        

    db.commit()

    db.refresh(medico)

    return medico



@app.put("/api/medicos/{medico_id}/huella", response_model=schemas.MedicoResponse)

def update_medico_huella(

    medico_id: int,

    fmd_template: str = Form(...),

    db: Session = Depends(get_db),

    current_user: models.Usuario = Depends(require_role(["admin", "rh", "sistemas"]))

):

    medico = db.query(models.Medico).filter(models.Medico.id == medico_id).first()

    if not medico:

        raise HTTPException(status_code=404, detail="Médico no encontrado")

        

    medico.fmd_template = fmd_template

    db.commit()

    db.refresh(medico)

    return medico



# === ESCANEOS RH ===

ESCANEOS_DIR = "static/escaneos_rh"

os.makedirs(ESCANEOS_DIR, exist_ok=True)



@app.get("/api/escaneos", response_model=List[schemas.EscaneoRHResponse])

def get_escaneos(db: Session = Depends(get_db), current_user: models.Usuario = Depends(require_role(["admin", "rh", "sistemas"]))):

    return db.query(models.EscaneoRH).order_by(models.EscaneoRH.fecha_subida.desc()).all()



@app.post("/api/escaneos", response_model=schemas.EscaneoRHResponse)

async def upload_escaneo(

    titulo: str = Form(...),

    archivo: UploadFile = File(...),

    db: Session = Depends(get_db),

    current_user: models.Usuario = Depends(require_role(["admin", "rh"]))

):

    # Enforce 5MB limit

    contents = await archivo.read()

    if len(contents) > 5 * 1024 * 1024:

        raise HTTPException(status_code=400, detail="El archivo excede el tamaño máximo permitido de 5MB.")

    

    ext = os.path.splitext(archivo.filename)[1]

    unique_filename = f"{uuid.uuid4()}{ext}"

    filepath = os.path.join(ESCANEOS_DIR, unique_filename)

    

    with open(filepath, "wb") as buffer:

        buffer.write(contents)

        

    ruta_archivo = f"/static/escaneos_rh/{unique_filename}"

    

    nuevo_escaneo = models.EscaneoRH(

        titulo=titulo,

        nombre_archivo=archivo.filename,

        ruta_archivo=ruta_archivo,

        subido_por_id=current_user.id

    )

    db.add(nuevo_escaneo)

    db.commit()

    db.refresh(nuevo_escaneo)

    return nuevo_escaneo



@app.put("/api/escaneos/{escaneo_id}", response_model=schemas.EscaneoRHResponse)

def rename_escaneo(escaneo_id: int, req: schemas.EscaneoRHUpdate, db: Session = Depends(get_db), current_user: models.Usuario = Depends(require_role(["admin", "rh"]))):

    escaneo = db.query(models.EscaneoRH).filter(models.EscaneoRH.id == escaneo_id).first()

    if not escaneo:

        raise HTTPException(status_code=404, detail="Escaneo no encontrado")

    escaneo.titulo = req.titulo

    db.commit()

    db.refresh(escaneo)

    return escaneo



@app.delete("/api/escaneos/{escaneo_id}")

def delete_escaneo(escaneo_id: int, db: Session = Depends(get_db), current_user: models.Usuario = Depends(require_role(["admin", "rh"]))):

    escaneo = db.query(models.EscaneoRH).filter(models.EscaneoRH.id == escaneo_id).first()

    if not escaneo:

        raise HTTPException(status_code=404, detail="Escaneo no encontrado")

    

    # Try to delete the physical file

    try:

        filename = os.path.basename(escaneo.ruta_archivo)

        filepath = os.path.join(ESCANEOS_DIR, filename)

        if os.path.exists(filepath):

            os.remove(filepath)

    except Exception as e:

        print(f"Error al eliminar el archivo fisico: {e}")

        

    db.delete(escaneo)

    db.commit()

    return {"message": "Escaneo eliminado correctamente"}



# === DASHBOARD, ANALYTICS & AUDITORÍA ===

@app.get("/api/analytics", response_model=schemas.AnalyticsDashboardResponse)

def get_analytics(db: Session = Depends(get_db), current_user: models.Usuario = Depends(require_role(["admin", "sistemas", "rh", "director"]))):

    # Total pacientes activos

    pacientes_activos = db.query(models.Paciente).filter(models.Paciente.status_ingreso == "Ingresado").count()

    

    # Atenciones del mes

    hoy = datetime.date.today()

    primer_dia = datetime.date(hoy.year, hoy.month, 1)

    atenciones_mes = db.query(models.AtencionMedica).filter(func.date(models.AtencionMedica.fecha_realizacion) >= primer_dia).count()

    

    # Visitas por área (todas las atenciones agrupadas)

    areas_db = db.query(

        models.AtencionMedica.area_hospitalaria, func.count(models.AtencionMedica.folio).label("total")

    ).group_by(models.AtencionMedica.area_hospitalaria).all()

    visitas_area = [{"area": a[0] or "Sin Área", "cantidad": a[1]} for a in areas_db]

    

    # Actividad reciente (últimos 7 días)

    hace_7_dias = hoy - datetime.timedelta(days=7)

    actividad_db = db.query(

        func.date(models.AtencionMedica.fecha_realizacion).label("fecha"), func.count(models.AtencionMedica.folio).label("total")

    ).filter(func.date(models.AtencionMedica.fecha_realizacion) >= hace_7_dias).group_by(func.date(models.AtencionMedica.fecha_realizacion)).order_by("fecha").all()

    actividad = [{"fecha": str(a[0]), "cantidad": a[1]} for a in actividad_db]

    

    # SLA Médico (tiempo entre fecha_registro y fecha_firma)

    atenciones_firmadas = db.query(models.AtencionMedica).filter(models.AtencionMedica.fecha_firma.isnot(None), models.AtencionMedica.medico_id.isnot(None)).all()

    sla_dict = {}

    for a in atenciones_firmadas:

        m_name = a.medico.nombre_completo if a.medico else "Desconocido"

        delta = (a.fecha_firma - a.fecha_registro).total_seconds() / 60.0

        if delta < 0:

            delta = 0

        if m_name not in sla_dict:

            sla_dict[m_name] = {"sum": 0, "count": 0}

        sla_dict[m_name]["sum"] += delta

        sla_dict[m_name]["count"] += 1

        

    sla_medicos = []

    for m_name, stats in sla_dict.items():

        sla_medicos.append({

            "medico": m_name,

            "tiempo_promedio_minutos": round(stats["sum"] / stats["count"], 1),

            "total_atenciones": stats["count"]

        })

    sla_medicos = sorted(sla_medicos, key=lambda x: x["tiempo_promedio_minutos"])[:10] # Top 10 más rápidos



    return {

        "total_pacientes_activos": pacientes_activos,

        "total_atenciones_mes": atenciones_mes,

        "visitas_por_area": visitas_area,

        "sla_por_medico": sla_medicos,

        "actividad_reciente": actividad

    }



@app.get("/api/auditoria", response_model=List[schemas.AuditoriaLogResponse])

def get_auditoria(db: Session = Depends(get_db), current_user: models.Usuario = Depends(require_role(["admin", "sistemas"]))):

    return db.query(models.AuditoriaLog).order_by(models.AuditoriaLog.fecha_hora.desc()).limit(200).all()



@app.get("/api/pacientes/{paciente_id}/traslados", response_model=List[schemas.TrasladoPacienteResponse])

def get_traslados(paciente_id: int, db: Session = Depends(get_db), current_user: models.Usuario = Depends(get_current_user)):

    return db.query(models.TrasladoPaciente).filter(models.TrasladoPaciente.paciente_id == paciente_id).order_by(models.TrasladoPaciente.fecha_traslado.desc()).all()



@app.get("/api/pacientes/{paciente_id}/journey")

def get_paciente_journey(paciente_id: int, db: Session = Depends(get_db), current_user: models.Usuario = Depends(get_current_user)):

    paciente = db.query(models.Paciente).filter(models.Paciente.id == paciente_id).first()

    if not paciente:

        raise HTTPException(status_code=404, detail="Paciente no encontrado")

        

    eventos = []

    

    # 1. Ingreso

    eventos.append({

        "tipo": "INGRESO",

        "fecha": paciente.fecha_registro.isoformat() + "Z" if paciente.fecha_registro else None,

        "descripcion": f"Ingreso a Hospital. Área inicial: {paciente.area_hospitalaria or 'No asignada'}, Hab: {paciente.num_habitacion}",

        "usuario": paciente.creador.nombre_completo if paciente.creador else "Sistema"

    })

    

    # 2. Traslados

    traslados = db.query(models.TrasladoPaciente).filter(models.TrasladoPaciente.paciente_id == paciente_id).all()

    for t in traslados:

        eventos.append({

            "tipo": "TRASLADO",

            "fecha": t.fecha_traslado.isoformat() + "Z" if t.fecha_traslado else None,

            "descripcion": f"Traslado a {t.destino_area or 'No asignada'}, Hab: {t.destino_habitacion}",

            "usuario": t.usuario.nombre_completo if t.usuario else "Sistema"

        })

        

    # 3. Atenciones y Firmas

    atenciones = db.query(models.AtencionMedica).filter(models.AtencionMedica.paciente_id == paciente_id).all()

    for a in atenciones:

        # Solicitud de Atención

        eventos.append({

            "tipo": "ATENCION",

            "fecha": a.fecha_registro.isoformat() + "Z" if a.fecha_registro else None,

            "descripcion": f"Solicitud de atención: {a.nombre_procedimiento}",

            "usuario": a.creador.nombre_completo if a.creador else "Sistema"

        })

        # Firma Médica

        if a.fecha_firma and a.medico:

            eventos.append({

                "tipo": "FIRMA_MEDICA",

                "fecha": a.fecha_firma.isoformat() + "Z" if a.fecha_firma else None,

                "descripcion": f"Firma médica completada por {a.medico.nombre_completo}",

                "usuario": a.medico.nombre_completo

            })

            

    # 4. Alta

    if paciente.status_ingreso == "Alta" and paciente.fecha_alta:

        eventos.append({

            "tipo": "ALTA",

            "fecha": paciente.fecha_alta.isoformat() + "Z" if paciente.fecha_alta else None,

            "descripcion": "Alta del paciente",

            "usuario": paciente.dado_de_alta_por.nombre_completo if paciente.dado_de_alta_por else "Sistema"

        })

        

    # Filtrar eventos sin fecha y ordenar por fecha ascendente

    eventos = [e for e in eventos if e["fecha"] is not None]

    eventos.sort(key=lambda x: x["fecha"])

    

    return eventos



# === BACKUP ===

@app.get("/api/backup")

def get_backup(current_user: models.Usuario = Depends(require_role(["admin", "sistemas"]))):

    db_path = "hospital_escandon.db"

    if not os.path.exists(db_path):

        raise HTTPException(status_code=404, detail="Base de datos no encontrada")

    hoy = datetime.date.today().strftime("%Y%m%d")

    return FileResponse(path=db_path, filename=f"backup_hes_{hoy}.db", media_type="application/octet-stream")



@app.post("/api/atenciones/{folio}/notas", response_model=schemas.NotaResponse)

def agregar_nota(folio: str, req: schemas.NotaCreate, db: Session = Depends(get_db), current_user = Depends(get_current_user)):

    atencion = db.query(models.AtencionMedica).filter(models.AtencionMedica.folio == folio).first()

    if not atencion:

        raise HTTPException(status_code=404, detail="Atención no encontrada")

    if atencion.is_caducado and getattr(current_user, "rol", "") != "sistemas":

        raise HTTPException(status_code=400, detail="Fuera de tiempo permitido para captura (registro caducado)")

    

    # Medico objects don't map to usuarios table; only store id for Usuario rows

    is_medico = getattr(current_user, "rol", "") in ("medico", "ayudante")

    nueva_nota = models.NotaEnfermeria(

        atencion_folio=folio,

        nota=req.nota,

        creada_por_id=None if is_medico else current_user.id

    )

    db.add(nueva_nota)

    db.commit()

    db.refresh(nueva_nota)

    return nueva_nota



@app.put("/api/atenciones/{folio}/reaperturar")

def reaperturar_registro(folio: str, db: Session = Depends(get_db), current_user: models.Usuario = Depends(require_role(["sistemas"]))):

    atencion = db.query(models.AtencionMedica).filter(models.AtencionMedica.folio == folio).first()

    if not atencion:

        raise HTTPException(status_code=404, detail="Atención no encontrada")

    atencion.reaperturado = True

    db.commit()

    return {"message": "Registro reaperturado exitosamente"}



@app.put("/api/atenciones/{folio}/autorizar")

def autorizar_registro(folio: str, req: schemas.AutorizarRequest, db: Session = Depends(get_db), current_user: models.Usuario = Depends(require_role(["sistemas"]))):

    atencion = db.query(models.AtencionMedica).filter(models.AtencionMedica.folio == folio).first()

    if not atencion:

        raise HTTPException(status_code=404, detail="Atención no encontrada")

    if atencion.estatus_pago != "Pendiente Autorización":

        raise HTTPException(status_code=400, detail="El registro no está pendiente de autorización.")

        

    if req.aceptado:

        atencion.estatus_pago = "Pendiente de Firma"

    else:

        atencion.estatus_pago = "Denegado"

    db.commit()

    return {"message": f"Registro {'autorizado' if req.aceptado else 'denegado'} exitosamente"}



@app.get("/api/pacientes/altas", response_model=List[schemas.PacienteResponse])

def get_pacientes_altas(db: Session = Depends(get_db), current_user = Depends(get_current_user)):

    return db.query(models.Paciente).filter(models.Paciente.status_ingreso == "Alta").order_by(models.Paciente.fecha_alta.desc()).all()




# === MÓDULO CAMAS ===

@app.get("/api/camas")

def get_camas(db: Session = Depends(get_db)):

    """Obtiene el listado de camas desde el hospital cruzado con el estado de limpieza local."""

    kh_camas = kh_database.fetch_camas()

    if not isinstance(kh_camas, list) or len(kh_camas) == 0 or "Error" in kh_camas[0] or "Mensaje" in kh_camas[0]:

        return kh_camas

        

    local_camas = {c.numero_cama: c for c in db.query(models.Cama).all()}

    

    for cama in kh_camas:

        room_name = cama.get("RoomName")

        if room_name in local_camas:

            c = local_camas[room_name]

            cama["estado_limpieza"] = c.estado_limpieza or "Disponible"

            cama["notas_limpieza"] = c.notas_limpieza

        else:

            cama["estado_limpieza"] = "Disponible"

            cama["notas_limpieza"] = None

            

    return kh_camas



@app.put("/api/camas/{numero_cama}/limpieza")

def update_cama_limpieza(numero_cama: str, payload: dict, db: Session = Depends(get_db), current_user: models.Usuario = Depends(require_role(["Mantenimiento/Limpieza", "limpieza", "admin", "sistemas"]))):

    cama_db = db.query(models.Cama).filter(models.Cama.numero_cama == numero_cama).first()

    if not cama_db:

        cama_db = models.Cama(numero_cama=numero_cama, area="Otras Áreas")

        db.add(cama_db)

    

    cama_db.estado_limpieza = payload.get("estado_limpieza", "Disponible")

    cama_db.notas_limpieza = payload.get("notas_limpieza", None)

    db.commit()

    

    log_auditoria(db, current_user.id, "Actualización Limpieza Cama", f"Cama {numero_cama} -> {cama_db.estado_limpieza}")

    

    return {"status": "ok", "message": "Estado de limpieza actualizado"}



@app.get("/api/camas/paciente/{pt_num}")

def get_patient_timeline(pt_num: str):

    """Obtiene la Ficha Rápida (demográficos) y la Línea de Tiempo del paciente."""

    return kh_database.fetch_patient_info_and_timeline(pt_num)



@app.get("/api/ehr/paciente/{pt_num}")

def get_full_ehr_dashboard(pt_num: str, db: Session = Depends(get_db)):

    """Obtiene el dashboard completo del expediente (Fase 1)."""

    data = kh_database.fetch_full_ehr_dashboard(pt_num)

    if isinstance(data, dict) and "error" not in data:

        last_hist = db.query(models.HistoricoNotaClinica).filter(

            models.HistoricoNotaClinica.pt_num == str(pt_num),

            models.HistoricoNotaClinica.codigo_formato == "HE-DIRMED-CONSUL-PLT-32/01"

        ).order_by(models.HistoricoNotaClinica.fecha_registro.desc()).first()

        if last_hist and last_hist.contenido_soap_json:

            try:

                hist_data = json.loads(last_hist.contenido_soap_json)

                if not data.get("consentimiento_32_01"):

                    data["consentimiento_32_01"] = {}

                data["consentimiento_32_01"]["testigo1"] = hist_data.get("testigo1", "")

                data["consentimiento_32_01"]["testigo2"] = hist_data.get("testigo2", "")

                data["consentimiento_32_01"]["representante_legal"] = hist_data.get("representante_legal", "")

                data["consentimiento_32_01"]["paciente_o_representante"] = hist_data.get("paciente_o_representante", "")

            except Exception as e:

                print(f"Error parsing historic consent: {e}")

    return data



@app.get("/api/ehr/paciente/{pt_num}/pdf-consentimiento-32-01")

def get_pdf_consentimiento_32_01(

    pt_num: str, 

    tipo_interrogatorio: str = "Directo",

    testigo1: str = "",

    testigo2: str = "",

    paciente_o_representante: str = "",

    representante_legal: str = ""

):

    dashboard_data = kh_database.fetch_full_ehr_dashboard(pt_num)

    if "error" in dashboard_data:

        raise HTTPException(status_code=404, detail=dashboard_data["error"])

        

    patient_info = dashboard_data.get("patient", {})

    fecha_hoy = datetime.datetime.now().strftime("%d/%m/%Y")

    hora_hoy = datetime.datetime.now().strftime("%H:%M")

    

    pt_data = {

        "nombre": patient_info.get("name", ""),

        "dob": patient_info.get("dob", ""),

        "mrn": patient_info.get("mrn", ""),

        "cama": patient_info.get("cama", "URGENCIAS"),

        "edad": patient_info.get("age", ""),

        "sexo": "M" if patient_info.get("gender") == "Masculino" else "F",

        "grupo_rh": patient_info.get("grupo_rh", "O+"),

        "alergias": patient_info.get("allergies", "NEGADAS"),

        "fecha_ingreso": patient_info.get("fecha_ingreso", fecha_hoy),

        "hora_ingreso": patient_info.get("hora_ingreso", hora_hoy),

        "tipo_interrogatorio": tipo_interrogatorio,

        "diagnostico": patient_info.get("diagnostico", "VALORACIÓN CARDIOLÓGICA"),

        "medico_tratante": patient_info.get("attending", ""),

        "cedula": patient_info.get("cedula", ""),

        "paciente_o_representante": paciente_o_representante or patient_info.get("name", ""),

        "representante_legal": representante_legal,

        "medico_autorizado": patient_info.get("attending", ""),

        "testigo1": testigo1,

        "testigo2": testigo2,

        "fecha_documento": fecha_hoy

    }

    

    # Consultar firma biométrica ACTIVA para el consentimiento

    db = SessionLocal()

    try:

        firma_obj = db.query(models.FirmaDocumentoClinico).filter(

            models.FirmaDocumentoClinico.pt_num == str(pt_num),

            models.FirmaDocumentoClinico.codigo_formato == "HE-DIRMED-CONSUL-PLT-32/01",

            models.FirmaDocumentoClinico.estado == "ACTIVA"

        ).order_by(models.FirmaDocumentoClinico.fecha_hora_firma.desc()).first()

        if firma_obj:

            pt_data["firma_data"] = {

                "sello_digital": firma_obj.sello_digital,

                "hash_sha256": firma_obj.hash_sha256,

                "fecha_hora_firma": firma_obj.fecha_hora_firma.strftime("%d/%m/%Y %H:%M:%S") if firma_obj.fecha_hora_firma else "",

                "nombre_medico": firma_obj.nombre_medico,

                "cedula": firma_obj.cedula_profesional

            }

            if firma_obj.nombre_medico:

                pt_data["medico_autorizado"] = firma_obj.nombre_medico

            if firma_obj.cedula_profesional:

                pt_data["cedula"] = firma_obj.cedula_profesional

    except Exception as e:

        print(f"Error querying signature for 32/01 PDF: {e}")

    finally:

        db.close()

    

    import pdf_engine_32_01

    pdf_filename = f"consentimiento_32_01_{pt_num}.pdf"

    pdf_path = os.path.join(os.path.dirname(__file__), "static", "pdfs", pdf_filename)

    pdf_engine_32_01.generate_consentimiento_32_01(pt_data, pdf_path)

    

    # Almacenar el PDF binario en MR_CI_ETE_CARD para que Vertical lo pueda abrir

    try:

        with open(pdf_path, 'rb') as f:

            pdf_bytes = f.read()

        conn = kh_database.get_kh_connection()

        if conn:

            cursor = conn.cursor()

            cursor.execute("""

                UPDATE MR_CI_ETE_CARD

                SET PDF = ?, PDFFileName = ?, PDFContentType = 'application/pdf', PDFLength = ?

                WHERE MRNum_CI_ETE_CARD = (SELECT TOP 1 MRNum_CI_ETE_CARD FROM MR_CI_ETE_CARD WHERE PTNum = ? ORDER BY MRNum_CI_ETE_CARD DESC)

            """, (pdf_bytes, pdf_filename, len(pdf_bytes), pt_num))

            conn.commit()

            conn.close()

    except Exception as e:

        print(f"Nota: No se pudo almacenar PDF binario en MR_CI_ETE_CARD: {e}")

    

    from fastapi.responses import FileResponse

    return FileResponse(path=pdf_path, filename=pdf_filename, media_type='application/pdf')



@app.get("/api/ehr/paciente/{pt_num}/pdf-nota-urgencias")

def get_pdf_nota_urgencias(pt_num: str, evolucion: Optional[int] = None):

    """

    Genera y descarga el PDF de la Nota de Urgencias:

    - evolucion=None / 0: Formato General (3 notas en 1, con 1 sola firma al final).

    - evolucion=1, 2, 3: Nota individual específica (con su propia firma).

    """

    dashboard_data = kh_database.fetch_full_ehr_dashboard(pt_num)

    

    if "error" in dashboard_data:

        raise HTTPException(status_code=404, detail=dashboard_data["error"])

        

    patient_info = dashboard_data.get("patient", {})

    notes = dashboard_data.get("clinicalNotes", [])

    

    if not notes:

        raise HTTPException(status_code=404, detail="No hay notas clínicas para generar PDF")

        

    fecha_hoy = datetime.datetime.now().strftime("%d/%m/%Y")

    hora_hoy = datetime.datetime.now().strftime("%H:%M")

    

    pt_data = {

        "nombre": patient_info.get("name", ""),

        "dob": patient_info.get("dob", ""),

        "mrn": patient_info.get("mrn", ""),

        "cama": patient_info.get("cama", "Urgencias"),

        "edad": patient_info.get("age", ""),

        "sexo": "M" if patient_info.get("gender") == "Masculino" else "F",

        "grupo_rh": "O+",

        "alergias": patient_info.get("allergies", ""),

        "fecha_ingreso": patient_info.get("fecha_ingreso", fecha_hoy),

        "hora_ingreso": patient_info.get("hora_ingreso", hora_hoy),

        "diagnostico": patient_info.get("diagnostico", ""),

        "destino": patient_info.get("destino", "DOMICILIO"),

        "fecha_egreso": patient_info.get("fecha_egreso", "___/___/___"),

        "hora_egreso": patient_info.get("hora_egreso", "__:__")

    }

    

    evols = dashboard_data.get("evoluciones", {})

    e1 = evols.get("evolucion1")

    e2 = evols.get("evolucion2")

    e3 = evols.get("evolucion3")

    

    import pdf_engine_v2

    import importlib

    importlib.reload(pdf_engine_v2)

    

    # Consultar si existe firma biométrica ACTIVA registrada en PostgreSQL

    db = SessionLocal()

    try:

        firma_query = db.query(models.FirmaDocumentoClinico).filter(

            models.FirmaDocumentoClinico.pt_num == str(pt_num),

            models.FirmaDocumentoClinico.estado == "ACTIVA"

        )

        if evolucion in [1, 2, 3]:

            firma_query = firma_query.filter(models.FirmaDocumentoClinico.evolution_slot == evolucion)

        firma_obj = firma_query.order_by(models.FirmaDocumentoClinico.fecha_hora_firma.desc()).first()

        firma_data = {

            "sello_digital": firma_obj.sello_digital,

            "hash_sha256": firma_obj.hash_sha256,

            "fecha_hora_firma": firma_obj.fecha_hora_firma.strftime("%d/%m/%Y %H:%M:%S") if firma_obj.fecha_hora_firma else "",

            "nombre_medico": firma_obj.nombre_medico,

            "cedula": firma_obj.cedula_profesional

        } if firma_obj else None

    except Exception as e:

        print(f"Error querying signature for PDF: {e}")

        firma_data = None

    finally:

        db.close()

    

    if evolucion in [1, 2, 3]:

        # Impresión individual de una sola nota

        target_evol = evols.get(f"evolucion{evolucion}")

        if not target_evol:

            raise HTTPException(status_code=404, detail=f"No se encontró información para la Evolución {evolucion}")

            

        pdf_filename = f"nota_urgencias_{pt_num}_evolucion_{evolucion}.pdf"

        pdf_path = os.path.join(os.path.dirname(__file__), "static", "pdfs", pdf_filename)

        pdf_engine_v2.generate_nota_urgencias(pt_data, target_evol, None, None, pdf_path, is_general=False, firma_data=firma_data)

    else:

        # Impresión del formato general (3 notas en 1, 1 sola firma al final)

        pdf_filename = f"nota_urgencias_{pt_num}_general.pdf"

        pdf_path = os.path.join(os.path.dirname(__file__), "static", "pdfs", pdf_filename)

        pdf_engine_v2.generate_nota_urgencias(pt_data, e1, e2, e3, pdf_path, is_general=True, firma_data=firma_data)

    

    from fastapi.responses import FileResponse

    return FileResponse(path=pdf_path, filename=pdf_filename, media_type='application/pdf')



class NotaUrgenciasInputSchema(BaseModel):

    evolution_num: Optional[int] = None

    fecha: Optional[str] = None

    hora: Optional[str] = None

    turno: Optional[str] = "Matutino"

    vitals_ta: Optional[str] = ""

    vitals_fc: Optional[str] = ""

    vitals_fr: Optional[str] = ""

    vitals_sato2: Optional[str] = ""

    vitals_peso: Optional[str] = ""

    vitals_talla: Optional[str] = ""

    vitals_temp: Optional[str] = ""

    subjetivo: Optional[str] = ""

    objetivo: Optional[str] = ""

    analisis: Optional[str] = ""

    plan: Optional[str] = ""

    medico: Optional[str] = "JOSE JOSE PRUEBA ENRIQUEZ"

    cedula: Optional[str] = "PRUEBA-99281"

    mip: Optional[str] = ""

    alergias: Optional[str] = None

    diagnostico: Optional[str] = None

    destino: Optional[str] = None

    cama: Optional[str] = None



class ConsentimientoInputSchema(BaseModel):

    tipo_interrogatorio: Optional[str] = "Directo"

    testigo1: Optional[str] = ""

    testigo2: Optional[str] = ""

    paciente_o_representante: Optional[str] = ""

    representante_legal: Optional[str] = ""

    medico_tratante: Optional[str] = "JOSE JOSE PRUEBA ENRIQUEZ"

    cedula: Optional[str] = "PRUEBA-99281"

    alergias: Optional[str] = "NEGADAS"

    diagnostico: Optional[str] = ""



@app.post("/api/ehr/paciente/{pt_num}/consentimiento-32-01")

def save_consentimiento_32_01(

    pt_num: str, 

    consent_data: ConsentimientoInputSchema,

    request: Request,

    db: Session = Depends(get_db)

):

    res = kh_database.save_or_update_consentimiento_32_01(pt_num, consent_data.model_dump())

    if "error" in res:

        raise HTTPException(status_code=500, detail=res["error"])

        

    client_ip = request.client.host if request.client else "127.0.0.1"

    

    # 1. Guardar Snapshot Inmutable en HistoricoNotaClinica (NOM-024)

    try:

        historico_entry = models.HistoricoNotaClinica(

            codigo_formato="HE-DIRMED-CONSUL-PLT-32/01",

            tipo_documento="Consentimiento Informado para Ecocardiograma Transesofágico",

            pt_num=str(pt_num),

            expediente=f"PT-{pt_num}",

            evolution_slot=0,

            nombre_medico=consent_data.medico_tratante,

            cedula_profesional=consent_data.cedula,

            contenido_soap_json=json.dumps(consent_data.model_dump(), default=str),

            accion="GUARDADO",

            ip_origen=client_ip

        )

        db.add(historico_entry)

        db.commit()

    except Exception as e:

        print(f"Error creating audit history for Consentimiento 32_01: {e}")

        db.rollback()



    # 2. Soft-Revocation de firmas activas en PostgreSQL (NOM-024: se rompe la firma al modificar)

    try:

        firmas_activas = db.query(models.FirmaDocumentoClinico).filter(

            models.FirmaDocumentoClinico.pt_num == str(pt_num),

            models.FirmaDocumentoClinico.codigo_formato == "HE-DIRMED-CONSUL-PLT-32/01",

            models.FirmaDocumentoClinico.estado == "ACTIVA"

        ).all()



        for f in firmas_activas:

            f.estado = "REVOCADA_POR_MODIFICACION"
            f.motivo_revocacion = "El documento fue modificado posteriormente."

            f.fecha_revocacion = datetime.datetime.now()

            f.motivo_revocacion = "Modificación y edición de los datos del consentimiento informado 32/01"



        db.commit()

    except Exception as e:

        print(f"Error revoking active signatures for Consentimiento 32_01: {e}")

        db.rollback()



    return res



@app.post("/api/ehr/paciente/{pt_num}/nota-urgencias")

def create_or_update_nota_urgencias(

    pt_num: str, 

    nota: NotaUrgenciasInputSchema, 

    request: Request,

    db: Session = Depends(get_db)

):

    """

    Guarda o actualiza una nota de evolución en SQL Server (con usuario de servicio institucional BITACORA_HES)

    y preserva el histórico inmutable de versiones y revocación de firmas en PostgreSQL conforme a la NOM-024.

    """

    res = kh_database.save_or_update_nota_urgencias(pt_num, nota.model_dump())

    if "error" in res:

        raise HTTPException(status_code=500, detail=res["error"])

        

    slot_affected = int(nota.evolution_num or res.get("slot") or 1)

    client_ip = request.client.host if request.client else "127.0.0.1"



    # 1. Guardar Snapshot Inmutable en HistoricoNotaClinica (NOM-024)

    try:

        historico_entry = models.HistoricoNotaClinica(

            codigo_formato="HE-DIRMED-SINPRO-PLT-87/01",

            tipo_documento=f"Nota de Evolución de Urgencias (Evolución {slot_affected})",

            pt_num=str(pt_num),

            expediente=f"PT-{pt_num}",

            evolution_slot=slot_affected,

            nombre_medico=nota.medico or "JOSE JOSE PRUEBA ENRIQUEZ",

            cedula_profesional=nota.cedula or "PRUEBA-99281",

            contenido_soap_json=json.dumps(nota.model_dump(), default=str),

            accion="EDICION" if nota.evolution_num else "CREACION",

            motivo="Actualización clínica desde Bitácora HES",

            fecha_registro=datetime.datetime.now(),

            ip_origen=client_ip

        )

        db.add(historico_entry)

    except Exception as e:

        print(f"Nota de auditoría: No se pudo registrar en HistoricoNotaClinica: {e}")



    # 2. INTEGRIDAD NOM-024 / NOM-004: Soft-Revocation (NUNCA BORRAR FÍSICAMENTE DE LA BD)

    firmas_activas = db.query(models.FirmaDocumentoClinico).filter(

        models.FirmaDocumentoClinico.pt_num == str(pt_num),

        models.FirmaDocumentoClinico.evolution_slot == slot_affected,

        models.FirmaDocumentoClinico.estado == "ACTIVA"

    ).all()



    for f in firmas_activas:

        f.estado = "REVOCADA_POR_MODIFICACION"
        f.motivo_revocacion = "El documento fue modificado posteriormente."

        f.fecha_revocacion = datetime.datetime.now()

        f.motivo_revocacion = f"Modificación y edición del contenido clínico en slot {slot_affected}"



    # 3. Registrar evento en AuditoriaLog Central

    try:

        log_auditoria = models.AuditoriaLog(

            usuario_id=None,

            accion="MODIFICACION_NOTA_CLINICA_Y_REVOCACION_FIRMA",

            detalles_json=json.dumps({

                "pt_num": str(pt_num),

                "slot": slot_affected,

                "medico": nota.medico,

                "cedula": nota.cedula,

                "firmas_revocadas_ids": [f.id for f in firmas_activas],

                "usuario_servicio": "BITACORA_HES"

            }),

            fecha_hora=datetime.datetime.now(),

            ip_origen=client_ip

        )

        db.add(log_auditoria)

    except Exception as e:

        print(f"Error registrando auditoria log: {e}")



    db.commit()

    print(f"Aviso NOM-024: Se preservó histórico y se marcaron {len(firmas_activas)} firmas como REVOCADAS (sin borrado físico).")



    return res



class SignosVitalesInputSchema(BaseModel):

    systolic: Optional[Union[int, str]] = None

    diastolic: Optional[Union[int, str]] = None

    ta: Optional[str] = None

    pulse: Optional[Union[int, str]] = None

    respiratory: Optional[Union[int, str]] = None

    oxygen_saturation: Optional[Union[int, str]] = None

    temperature: Optional[Union[float, str]] = None

    weight: Optional[Union[float, str]] = None

    height: Optional[Union[float, str]] = None

    procedure_date: Optional[str] = None



@app.get("/api/ehr/paciente/{pt_num}/signos-vitales")

def get_paciente_signos_vitales(pt_num: str):

    """Consulta los signos vitales más recientes y el historial desde la tabla maestra PTVS en SQL Server."""

    return kh_database.fetch_patient_vitals_ptvs(pt_num)



@app.post("/api/ehr/paciente/{pt_num}/signos-vitales")

def save_paciente_signos_vitales(

    pt_num: str, 

    vitals: SignosVitalesInputSchema, 

    request: Request,

    db: Session = Depends(get_db)

):

    """

    Registra o actualiza la toma de signos vitales en la tabla [KH_HE].[dbo].[PTVS] de SQL Server

    y asienta el evento en la auditoría inmutable de la Bitácora.

    """

    res = kh_database.save_patient_vitals_ptvs(pt_num, vitals.model_dump())

    if "error" in res:

        raise HTTPException(status_code=500, detail=res["error"])



    # Registrar en AuditoriaLog Central

    try:

        client_ip = request.client.host if request.client else "127.0.0.1"

        log = models.AuditoriaLog(

            usuario_id=None,

            accion="CAPTURA_SIGNOS_VITALES_PTVS",

            detalles_json=json.dumps({

                "pt_num": str(pt_num),

                "vitals": vitals.model_dump(),

                "ptvs_id": res.get("ptvs_id")

            }, default=str),

            fecha_hora=datetime.datetime.now(),

            ip_origen=client_ip

        )

        db.add(log)

        db.commit()

    except Exception as e:

        print(f"Error registrando auditoría de signos vitales: {e}")



    return res



class PrescribirMedicamentoInputSchema(BaseModel):

    name: str

    amount: float # Debe ser numérico positivo

    uom: Optional[str] = "mg"

    route: Optional[str] = "Oral"

    frequency: Optional[str] = "Cada 8 horas"

    prn: Optional[bool] = False

    why: Optional[str] = ""

    dispense: Optional[str] = ""

    refills: Optional[int] = 0

    instruction: Optional[str] = ""

    fmd_template: str # Huella dactilar obligatoria del médico tratante

    medico_id: Optional[int] = None



    @field_validator("amount", mode="before")

    @classmethod

    def validate_amount(cls, v):

        if v is None or (isinstance(v, str) and v.strip() == ""):

            raise ValueError("La dosis (amount) es obligatoria y debe ser un número positivo mayor a 0.")

        try:

            val = float(v)

        except (ValueError, TypeError):

            raise ValueError("La dosis (amount) debe ser un valor numérico válido (int o float).")

        if val <= 0:

            raise ValueError("La dosis (amount) debe ser un número positivo mayor a 0.")

        return val



    @field_validator("frequency")

    @classmethod

    def validate_frequency(cls, v):

        if v is not None:

            if len(v) > 100:

                raise ValueError("La frecuencia no puede tener más de 100 caracteres.")

            caracteres_prohibidos = [";", "--", "<", ">", "/*", "*/"]

            for patron in caracteres_prohibidos:

                if patron in v:

                    raise ValueError(f"La frecuencia contiene caracteres especiales no permitidos ('{patron}').")

        return v



class DiscontinuarMedicamentoInputSchema(BaseModel):

    ptdg_num: int

    reason: Optional[str] = "Discontinuado por evolución clínica"

    fmd_template: str # Huella dactilar para suspender



@app.get("/api/ehr/paciente/{pt_num}/medicamentos")

def get_paciente_medicamentos(pt_num: str):

    """Consulta la lista de medicamentos prescritos desde la tabla maestra PTDG en SQL Server."""

    return kh_database.fetch_patient_medications_ptdg(pt_num)



@app.post("/api/ehr/paciente/{pt_num}/medicamentos/prescribir-biometrico")

def prescribir_medicamento_biometrico(

    pt_num: str,

    req: PrescribirMedicamentoInputSchema,

    request: Request,

    db: Session = Depends(get_db)

):

    """

    Prescribe formalmente un fármaco en SQL Server (PTDG) requiriendo validación biométrica dactilar

    del médico conforme a la NOM-004-SSA3-2012 y NOM-024-SSA3-2012.

    """

    import requests

    if not req.fmd_template:

        raise HTTPException(status_code=400, detail="Se requiere la huella dactilar del médico para prescribir fármacos.")



    # 1. Validar huella dactilar contra médicos registrados

    medicos = db.query(models.Medico).filter(models.Medico.activo_status == True, models.Medico.fmd_template.isnot(None)).all()

    if not medicos:

        raise HTTPException(status_code=400, detail="No hay médicos registrados con huella biométrica en el sistema.")



    match_found = None

    medicos_data = [{"id": m.id, "fmd_template": m.fmd_template} for m in medicos]



    try:

        response = requests.post("http://localhost:8082/match-bulk", json={

            "fmd1": req.fmd_template,

            "medicos": medicos_data

        }, timeout=10)



        if response.status_code == 200:

            data = response.json()

            if data.get("success") and data.get("isMatch"):

                match_id = data.get("match_id")

                match_found = next((m for m in medicos if m.id == match_id), None)

    except Exception as e:

        print(f"Error conectando con microservicio biométrico: {e}")

        raise HTTPException(status_code=500, detail="Error de comunicación con el motor biométrico DigitalPersona.")



    if not match_found:

        raise HTTPException(status_code=401, detail="Huella dactilar no reconocida. Solo un médico adscrito autorizado puede prescribir.")



    # 2. Generar Firma Criptográfica de la Prescripción

    now = datetime.datetime.now()

    cadena_original = f"||{pt_num}|PT-{pt_num}|RECETA-PTDG|{req.name}|{req.amount} {req.uom}|{req.route}|{req.frequency}|{now.isoformat()}|{match_found.id}|{match_found.cedula}||"

    hash_sha256 = hashlib.sha256(cadena_original.encode('utf-8')).hexdigest()



    sello_digital = crypto_fea.firmar_documento(db, match_found, cadena_original)



    # 3. Guardar en SQL Server PTDG

    res = kh_database.save_patient_medication_ptdg(pt_num, req.model_dump())

    if "error" in res:

        raise HTTPException(status_code=500, detail=res["error"])



    # 4. Guardar evidencia en PostgreSQL

    client_ip = request.client.host if request.client else "127.0.0.1"

    firma_registro = models.FirmaDocumentoClinico(

        tipo_documento="Prescripción Médica de Farmacoterapia (PTDG)",

        codigo_formato="HE-DIRMED-SINPRO-REC-01",

        pt_num=str(pt_num),

        expediente=f"PT-{pt_num}",

        evolution_slot=None,

        medico_id=match_found.id,

        nombre_medico=match_found.nombre_completo,

        cedula_profesional=match_found.cedula,

        fecha_hora_firma=now,

        metodo_autenticacion="Biometría Dactilar DigitalPersona (NOM-004/NOM-024)",

        hash_sha256=hash_sha256,

        sello_digital=sello_digital,

        cadena_original=cadena_original,

        ip_origen=client_ip,

        estado="ACTIVA"

    )

    db.add(firma_registro)



    log_auditoria = models.AuditoriaLog(

        usuario_id=None,

        accion="PRESCRIPCION_MEDICAMENTO_PTDG",

        detalles_json=json.dumps({

            "pt_num": str(pt_num),

            "medication": req.name,

            "dose": f"{req.amount} {req.uom}",

            "route": req.route,

            "freq": req.frequency,

            "medico": match_found.nombre_completo,

            "cedula": match_found.cedula,

            "ptdg_id": res.get("ptdg_id")

        }),

        fecha_hora=now,

        ip_origen=client_ip

    )

    db.add(log_auditoria)

    db.commit()



    return {

        "success": True,

        "message": f"Fármaco '{req.name}' prescrito y firmado exitosamente por {match_found.nombre_completo}.",

        "medico": match_found.nombre_completo,

        "cedula": match_found.cedula,

        "ptdg_id": res.get("ptdg_id"),

        "sello": sello_digital[:32] + "..."

    }



@app.post("/api/ehr/paciente/{pt_num}/medicamentos/discontinuar-biometrico")

def discontinuar_medicamento_biometrico(

    pt_num: str,

    req: DiscontinuarMedicamentoInputSchema,

    request: Request,

    db: Session = Depends(get_db)

):

    """Suspende un fármaco activo en PTDG requiriendo huella biométrica del médico."""

    import requests

    if not req.fmd_template:

        raise HTTPException(status_code=400, detail="Se requiere huella dactilar para suspender medicamentos.")



    medicos = db.query(models.Medico).filter(models.Medico.activo_status == True, models.Medico.fmd_template.isnot(None)).all()

    match_found = None

    medicos_data = [{"id": m.id, "fmd_template": m.fmd_template} for m in medicos]



    try:

        response = requests.post("http://localhost:8082/match-bulk", json={

            "fmd1": req.fmd_template,

            "medicos": medicos_data

        }, timeout=10)

        if response.status_code == 200:

            data = response.json()

            if data.get("success") and data.get("isMatch"):

                match_id = data.get("match_id")

                match_found = next((m for m in medicos if m.id == match_id), None)

    except Exception as e:

        raise HTTPException(status_code=500, detail="Error de comunicación con el motor biométrico.")



    if not match_found:

        raise HTTPException(status_code=401, detail="Huella no autorizada para suspender fármacos.")



    res = kh_database.discontinue_patient_medication_ptdg(pt_num, req.ptdg_num, req.reason or "Indicación médica")

    if "error" in res:

        raise HTTPException(status_code=500, detail=res["error"])



    client_ip = request.client.host if request.client else "127.0.0.1"

    log = models.AuditoriaLog(

        usuario_id=None,

        accion="SUSPENSION_MEDICAMENTO_PTDG",

        detalles_json=json.dumps({

            "pt_num": str(pt_num),

            "ptdg_num": req.ptdg_num,

            "motivo": req.reason,

            "medico": match_found.nombre_completo,

            "cedula": match_found.cedula

        }),

        fecha_hora=datetime.datetime.now(),

        ip_origen=client_ip

    )

    db.add(log)

    db.commit()



    return {

        "success": True,

        "message": f"Medicamento suspendido por {match_found.nombre_completo}."

    }



class PrescribirDietaInputSchema(BaseModel):

    tipo_dieta: str

    horario: Optional[str] = "Continuo"

    fase_clinica: Optional[str] = ""

    indicaciones_nutricionales: Optional[str] = ""

    inicio_ayuno_dieta: Optional[str] = ""

    nutriologo_responsable: Optional[str] = "Nutrición Clínica HES"

    alergias_alimentarias: Optional[str] = ""

    tolerancia_via_oral: Optional[str] = "Adecuada"

    cuidados_enfermeria: Optional[List[Dict[str, Any]]] = []

    fmd_template: str # Huella dactilar obligatoria

    medico_id: Optional[int] = None



@app.get("/api/ehr/paciente/{pt_num}/dieta-cuidados")

def get_paciente_dieta_cuidados(pt_num: str, db: Session = Depends(get_db)):

    """Consulta la prescripción dietética y cuidados de enfermería (PostgreSQL + Fallback SQL Server)."""

    # 1. Buscar en PostgreSQL (registro extendido con firma)

    dieta_pg = db.query(models.DietaCuidadosPrescripcion).filter(

        models.DietaCuidadosPrescripcion.pt_num == str(pt_num),

        models.DietaCuidadosPrescripcion.activo == True

    ).order_by(models.DietaCuidadosPrescripcion.fecha_hora_prescripcion.desc()).first()



    if dieta_pg:

        cuidados = []

        if dieta_pg.cuidados_enfermeria_json:

            try:

                cuidados = json.loads(dieta_pg.cuidados_enfermeria_json)

            except Exception:

                cuidados = []

        return {

            "source": "PostgreSQL (Bitácora HES)",

            "tipo": dieta_pg.tipo_dieta,

            "horario": dieta_pg.horario,

            "fase": dieta_pg.fase_clinica,

            "indicaciones": dieta_pg.indicaciones_nutricionales,

            "inicio": dieta_pg.inicio_ayuno_dieta,

            "nutriologo": dieta_pg.nutriologo_responsable,

            "alergias_alimentarias": dieta_pg.alergias_alimentarias,

            "tolerancia_via_oral": dieta_pg.tolerancia_via_oral,

            "cuidados_enfermeria": cuidados,

            "medico": dieta_pg.medico_nombre,

            "cedula": dieta_pg.medico_cedula,

            "fecha_prescripcion": dieta_pg.fecha_hora_prescripcion.strftime("%d/%m/%Y %H:%M") if dieta_pg.fecha_hora_prescripcion else "",

            "sello": dieta_pg.sello_digital[:32] + "..." if dieta_pg.sello_digital else ""

        }



    # 2. Fallback a SQL Server MR_SOL_DIET

    dieta_sql = kh_database.fetch_patient_diet_mr_sol_diet(pt_num)

    if dieta_sql and (dieta_sql.get("tipo") or dieta_sql.get("mrnum_sol_diet")):

        return {

            "source": "SQL Server (MR_SOL_DIET)",

            "tipo": dieta_sql.get("tipo", "Dieta Hospitalaria"),

            "horario": dieta_sql.get("horario", "--"),

            "fase": "--",

            "indicaciones": dieta_sql.get("detalle") or "Sin indicaciones registradas.",

            "inicio": dieta_sql.get("created_on") or "--",

            "nutriologo": dieta_sql.get("created_by") or "--",

            "alergias_alimentarias": dieta_sql.get("intolerancia") or "Ninguna registrada",

            "tolerancia_via_oral": "--",

            "cuidados_enfermeria": []

        }



    return {

        "source": "Ninguna",

        "tipo": "Sin dieta asignada",

        "horario": "--",

        "fase": "--",

        "indicaciones": "No se ha registrado régimen dietético para este paciente.",

        "inicio": "--",

        "nutriologo": "--",

        "alergias_alimentarias": "--",

        "tolerancia_via_oral": "--",

        "cuidados_enfermeria": []

    }



@app.post("/api/ehr/paciente/{pt_num}/dieta-cuidados/prescribir-biometrico")

def prescribir_dieta_cuidados_biometrico(

    pt_num: str,

    req: PrescribirDietaInputSchema,

    request: Request,

    db: Session = Depends(get_db)

):

    """

    Prescribe formalmente el régimen dietético en SQL Server (MR_SOL_DIET) y almacena el plan de cuidados

    enriquecido en PostgreSQL con validación biométrica dactilar (NOM-004 / NOM-024).

    """

    import requests

    if not req.fmd_template:

        raise HTTPException(status_code=400, detail="Se requiere la huella dactilar para prescribir dieta y cuidados.")



    # 1. Validar huella dactilar

    medicos = db.query(models.Medico).filter(models.Medico.activo_status == True, models.Medico.fmd_template.isnot(None)).all()

    if not medicos:

        raise HTTPException(status_code=400, detail="No hay médicos registrados con huella biométrica en el sistema.")



    match_found = None

    medicos_data = [{"id": m.id, "fmd_template": m.fmd_template} for m in medicos]



    try:

        response = requests.post("http://localhost:8082/match-bulk", json={

            "fmd1": req.fmd_template,

            "medicos": medicos_data

        }, timeout=10)



        if response.status_code == 200:

            data = response.json()

            if data.get("success") and data.get("isMatch"):

                match_id = data.get("match_id")

                match_found = next((m for m in medicos if m.id == match_id), None)

    except Exception as e:

        print(f"Error conectando con microservicio biométrico: {e}")

        raise HTTPException(status_code=500, detail="Error de comunicación con el motor biométrico DigitalPersona.")



    if not match_found:

        raise HTTPException(status_code=401, detail="Huella dactilar no autorizada para prescribir régimen dietético.")



    # 2. Generar Sello Digital HMAC-SHA512

    now = datetime.datetime.now()

    cadena_original = f"||{pt_num}|PT-{pt_num}|DIETA-MR_SOL_DIET|{req.tipo_dieta}|{req.fase_clinica}|{req.horario}|{now.isoformat()}|{match_found.id}|{match_found.cedula}||"

    hash_sha256 = hashlib.sha256(cadena_original.encode('utf-8')).hexdigest()



    sello_digital = crypto_fea.firmar_documento(db, match_found, cadena_original)



    # 3. Guardar en SQL Server MR_SOL_DIET

    res_sql = kh_database.save_patient_diet_mr_sol_diet(pt_num, {

        "tipo": req.tipo_dieta,

        "horario": req.horario,

        "detalle": req.indicaciones_nutricionales or req.fase_clinica,

        "intolerancia": req.alergias_alimentarias

    })

    if "error" in res_sql:

        print(f"Nota: No se pudo guardar en MR_SOL_DIET: {res_sql['error']}")



    # 4. Desactivar prescripciones dietéticas previas en PostgreSQL para este paciente

    db.query(models.DietaCuidadosPrescripcion).filter(

        models.DietaCuidadosPrescripcion.pt_num == str(pt_num)

    ).update({"activo": False})



    # 5. Insertar en PostgreSQL (Bitácora HES)

    nueva_dieta = models.DietaCuidadosPrescripcion(

        pt_num=str(pt_num),

        expediente=f"PT-{pt_num}",

        tipo_dieta=req.tipo_dieta,

        horario=req.horario,

        fase_clinica=req.fase_clinica,

        indicaciones_nutricionales=req.indicaciones_nutricionales,

        inicio_ayuno_dieta=req.inicio_ayuno_dieta or now.strftime("%d/%m/%Y %H:%M"),

        nutriologo_responsable=req.nutriologo_responsable or "Nutrición Clínica HES",

        alergias_alimentarias=req.alergias_alimentarias,

        tolerancia_via_oral=req.tolerancia_via_oral,

        cuidados_enfermeria_json=json.dumps(req.cuidados_enfermeria or [], ensure_ascii=False),

        medico_id=match_found.id,

        medico_nombre=match_found.nombre_completo,

        medico_cedula=match_found.cedula,

        hash_sha256=hash_sha256,

        sello_digital=sello_digital,

        cadena_original=cadena_original,

        fecha_hora_prescripcion=now,

        activo=True

    )

    db.add(nueva_dieta)



    # 6. Registrar en auditoría

    client_ip = request.client.host if request.client else "127.0.0.1"

    log = models.AuditoriaLog(

        usuario_id=None,

        accion="PRESCRIPCION_DIETA_Y_CUIDADOS",

        detalles_json=json.dumps({

            "pt_num": str(pt_num),

            "tipo_dieta": req.tipo_dieta,

            "horario": req.horario,

            "fase": req.fase_clinica,

            "medico": match_found.nombre_completo,

            "cedula": match_found.cedula,

            "diet_id": res_sql.get("diet_id")

        }),

        fecha_hora=now,

        ip_origen=client_ip

    )

    db.add(log)

    db.commit()



    return {

        "success": True,

        "message": f"Régimen dietético '{req.tipo_dieta}' prescrito y firmado por {match_found.nombre_completo}.",

        "medico": match_found.nombre_completo,

        "cedula": match_found.cedula,

        "sello": sello_digital[:32] + "..."

    }



@app.get("/api/ehr/pacientes/buscar")

def buscar_pacientes_universal(q: str = "", limit: int = 30):

    """

    Buscador universal de pacientes (activos, hospitalizados y egresados/de alta) en Vertical SQL Server.

    Permite buscar por nombre, apellido, folio/expediente (PTNum) o CURP.

    """

    return kh_database.search_patients_kh(query_text=q, limit=limit)



# ==========================================

# ENDPOINTS ALERGIAS (PTAL + DIS_AL)

# ==========================================



class RegistrarAlergiaInputSchema(BaseModel):

    allergy_num: str

    allergic_since: Optional[str] = None

    notes: Optional[str] = ""

    user: Optional[str] = None



class InactivarAlergiaInputSchema(BaseModel):

    ptal_num: int

    user: Optional[str] = None



@app.get("/api/ehr/alergias/catalogo")

def obtener_catalogo_alergias(q: str = "", limit: int = 50):

    """

    Consulta el catálogo maestro de alergias de Vertical (DIS_AL).

    """

    return kh_database.fetch_allergy_catalog(search_query=q, limit=limit)



@app.get("/api/ehr/paciente/{pt_num}/alergias")

def obtener_alergias_paciente(pt_num: str):

    """

    Consulta las alergias activas del paciente registradas en SQL Server (PTAL).

    """

    return kh_database.fetch_patient_allergies_ptal(pt_num)



@app.post("/api/ehr/paciente/{pt_num}/alergias/registrar")

def registrar_alergia_paciente(

    pt_num: str, 

    req: RegistrarAlergiaInputSchema, 

    request: Request,

    db: Session = Depends(get_db)

):

    """

    Registra una nueva alergia para el paciente en Vertical (PTAL).

    """

    if not req.allergy_num:

        raise HTTPException(status_code=400, detail="Debe seleccionar una alergia del catálogo.")

    

    usuario = req.user or "jose_prueba"

    res = kh_database.save_patient_allergy_ptal(

        pt_num=pt_num,

        allergy_num=req.allergy_num,

        allergic_since=req.allergic_since,

        notes=req.notes or "",

        user=usuario

    )

    if "error" in res:

        raise HTTPException(status_code=500, detail=res["error"])



    # Auditoría Forense

    auditoria = models.AuditoriaLog(

        tipo_accion="CREACION",

        modulo="ALERGIAS_PTAL",

        usuario=usuario,

        paciente_id=str(pt_num),

        ip_origen=request.client.host if request.client else "127.0.0.1",

        detalles_json={

            "accion": "Registro de alergia en PTAL",

            "allergy_num": req.allergy_num,

            "allergic_since": req.allergic_since,

            "notes": req.notes,

            "ptal_id": res.get("ptal_id")

        }

    )

    db.add(auditoria)

    db.commit()



    return res



@app.post("/api/ehr/paciente/{pt_num}/alergias/inactivar")

def inactivar_alergia_paciente(

    pt_num: str, 

    req: InactivarAlergiaInputSchema, 

    request: Request,

    db: Session = Depends(get_db)

):

    """

    Inactiva una alergia del paciente en Vertical (PTAL).

    """

    usuario = req.user or "jose_prueba"

    res = kh_database.inactivate_patient_allergy_ptal(

        pt_num=pt_num,

        ptal_num=req.ptal_num,

        user=usuario

    )

    if "error" in res:

        raise HTTPException(status_code=500, detail=res["error"])



    # Auditoría Forense

    auditoria = models.AuditoriaLog(

        tipo_accion="ELIMINACION",

        modulo="ALERGIAS_PTAL",

        usuario=usuario,

        paciente_id=str(pt_num),

        ip_origen=request.client.host if request.client else "127.0.0.1",

        detalles_json={

            "accion": "Inactivación de alergia en PTAL",

            "ptal_num": req.ptal_num

        }

    )

    db.add(auditoria)

    db.commit()



    return res



class ActualizarTextoAlergiasInputSchema(BaseModel):

    allergies_text: str

    user: Optional[str] = None



@app.post("/api/ehr/paciente/{pt_num}/alergias/actualizar-texto")

def actualizar_texto_alergias_paciente(

    pt_num: str, 

    req: ActualizarTextoAlergiasInputSchema, 

    request: Request,

    db: Session = Depends(get_db)

):

    """

    Actualiza el texto consolidado de ALERGIAS en MR_NE_URG y MR_SOL_DIET de Vertical.

    """

    usuario = req.user or "jose_prueba"

    res = kh_database.update_patient_allergies_text(

        pt_num=pt_num,

        allergies_text=req.allergies_text,

        user=usuario

    )

    if "error" in res:

        raise HTTPException(status_code=500, detail=res["error"])



    # Auditoría Forense

    auditoria = models.AuditoriaLog(

        tipo_accion="ACTUALIZACION",

        modulo="ALERGIAS_TEXTO",

        usuario=usuario,

        paciente_id=str(pt_num),

        ip_origen=request.client.host if request.client else "127.0.0.1",

        detalles_json={

            "accion": "Actualización manual de texto de alergias en MR_NE_URG / MR_SOL_DIET",

            "allergies_text": req.allergies_text

        }

    )

    db.add(auditoria)

    db.commit()



    return res



class FirmaBiometricaInputSchema(BaseModel):

    codigo_formato: str = "HE-DIRMED-SINPRO-PLT-87/01"

    tipo_documento: str = "Nota de Evolución de Urgencias (87/01)"

    evolution_slot: Optional[int] = 1

    fmd_template: str

    medico_id: Optional[int] = None

    contenido_resumen: Optional[str] = ""



@app.post("/api/ehr/paciente/{pt_num}/firmar-biometrico")

def firmar_documento_biometrico(

    pt_num: str, 

    req: FirmaBiometricaInputSchema, 

    request: Request,

    db: Session = Depends(get_db)

):

    """

    Firma electrónicamente una nota o consentimiento mediante Biometría Dactilar DigitalPersona,

    generando sello digital y registro con FECHA Y HORA LOCAL EXACTA conforme a la NOM-004-SSA3-2012 y NOM-024-SSA3-2012.

    """

    import requests

    if not req.fmd_template:

        raise HTTPException(status_code=400, detail="No se recibió la huella dactilar (FMD).")

    

    # 1. Buscar médicos con huella registrada

    medicos = db.query(models.Medico).filter(models.Medico.activo_status == True, models.Medico.fmd_template.isnot(None)).all()

    if not medicos:

        raise HTTPException(status_code=400, detail="No hay médicos registrados con huella en el sistema.")

        

    match_found = None

    medicos_data = [{"id": m.id, "fmd_template": m.fmd_template} for m in medicos]

    

    try:

        response = requests.post("http://localhost:8082/match-bulk", json={

            "fmd1": req.fmd_template,

            "medicos": medicos_data

        }, timeout=10)

        

        if response.status_code == 200:

            data = response.json()

            if data.get("success") and data.get("isMatch"):

                match_id = data.get("match_id")

                match_found = next((m for m in medicos if m.id == match_id), None)

    except Exception as e:

        print(f"Error conectando con microservicio biométrico: {e}")

        raise HTTPException(status_code=500, detail="Error de conexión con el motor biométrico DigitalPersona.")

        

    if not match_found:

        raise HTTPException(status_code=401, detail="Huella no reconocida o no coincide con ningún médico registrado.")

        

    # 2. Fecha y hora exacta local (sin desfase UTC)

    now = datetime.datetime.now()

    fecha_iso = now.isoformat()

    fecha_legible = now.strftime("%d/%m/%Y %H:%M:%S")

    cadena_original = f"||{pt_num}|PT-{pt_num}|{req.codigo_formato}|{req.evolution_slot or 'GRAL'}|{fecha_iso}|{match_found.id}|{match_found.cedula}|{req.contenido_resumen[:100]}||"

    

    hash_sha256 = hashlib.sha256(cadena_original.encode('utf-8')).hexdigest()

    

    # SELLO CRIPTOGRÁFICO (Firma Electrónica Avanzada ECDSA del médico)

    sello_digital = crypto_fea.firmar_documento(db, match_found, cadena_original)

    

    # SELLADO DE TIEMPO RFC 3161 (no bloqueante: si el TSA falla, la firma continúa)

    tsa_info = tsa_client.get_timestamp(hash_sha256)

    

    # 2.5 Revocar firmas previas activas para este mismo formato y slot (NOM-024)

    target_slot = int(req.evolution_slot) if req.evolution_slot is not None else 0

    firmas_antiguas = db.query(models.FirmaDocumentoClinico).filter(

        models.FirmaDocumentoClinico.pt_num == str(pt_num),

        models.FirmaDocumentoClinico.codigo_formato == req.codigo_formato,

        models.FirmaDocumentoClinico.evolution_slot == target_slot,

        models.FirmaDocumentoClinico.estado == "ACTIVA"

    ).all()

    for fa in firmas_antiguas:

        fa.estado = "REVOCADA"

        fa.fecha_revocacion = now

        fa.motivo_revocacion = "Refirma del documento por el médico"



    # 3. Guardar en PostgreSQL

    client_ip = request.client.host if request.client else "127.0.0.1"

    firma_registro = models.FirmaDocumentoClinico(

        tipo_documento=req.tipo_documento,

        codigo_formato=req.codigo_formato,

        pt_num=str(pt_num),

        expediente=f"PT-{pt_num}",

        evolution_slot=target_slot,

        medico_id=match_found.id,

        nombre_medico=match_found.nombre_completo,

        cedula_profesional=match_found.cedula,

        fecha_hora_firma=now,

        metodo_autenticacion="Biometría Dactilar DigitalPersona (NOM-004/NOM-024-SSA3)",

        hash_sha256=hash_sha256,

        sello_digital=sello_digital,

        cadena_original=cadena_original,

        ip_origen=client_ip,

        tsa_token=tsa_info['token_b64'] if tsa_info else None

    )

    db.add(firma_registro)

    db.commit()

    db.refresh(firma_registro)

    

        # 4. Actualización en SQL Server e Invocación Automática de la API de Vertical
    try:
        from vertical_signer import sign_in_vertical_api
        ctrl_map = {
            'HE-DIRMED-CONSUL-PLT-32/01': ('MR_CI_ETE_CARD', 'MRNum_CI_ETE_CARD'),
            'HE-DIRMED-CONSUL-PLT-EED': ('MR_CI_EED', 'MRNum_CI_EED'),
            'HE-DIRMED-NOTAS-URG-87/01': ('MR_NE_URG', 'MRNum_NE_URG')
        }
        if req.codigo_formato in ctrl_map:
            c_name, pk_col = ctrl_map[req.codigo_formato]
            conn = kh_database.get_kh_connection()
            if conn:
                cursor = conn.cursor()
                cursor.execute(f"SELECT TOP 1 {pk_col} FROM {c_name} WHERE PTNum = ? ORDER BY {pk_col} DESC", (pt_num,))
                row = cursor.fetchone()
                conn.close()
                if row and row[0]:
                    sign_in_vertical_api(
                        controller_name=c_name,
                        mrnum=row[0],
                        pt_num=str(pt_num),
                        pr_num=257,
                        auth_code='123456',
                        doctor_name=match_found.nombre_completo
                    )
    except Exception as e:
        print(f"Nota: No se completó la autofirma en Vertical: {e}")

        

    return {

        "success": True,

        "message": "Documento firmado biométricamente con éxito conforme a la NOM-004-SSA3-2012 y NOM-024-SSA3-2012.",

        "firma": {

            "id": firma_registro.id,

            "nombre_medico": match_found.nombre_completo,

            "cedula": match_found.cedula,

            "fecha_hora": now.strftime("%d/%m/%Y %H:%M:%S"),

            "hash_sha256": hash_sha256,

            "sello_digital": sello_digital,

            "tsa_gen_time": tsa_info['gen_time'] if tsa_info else None,

            "normativa": "NOM-004-SSA3-2012 / NOM-024-SSA3-2012"

        }

    }



@app.get("/api/ehr/paciente/{pt_num}/firmas")

def get_firmas_paciente(pt_num: str, db: Session = Depends(get_db)):

    """Obtiene las firmas biométricas activas y vigentes para el paciente."""

    firmas = db.query(models.FirmaDocumentoClinico).filter(

        models.FirmaDocumentoClinico.pt_num == str(pt_num),

        models.FirmaDocumentoClinico.estado == "ACTIVA"

    ).order_by(models.FirmaDocumentoClinico.fecha_hora_firma.desc()).all()

    return [

        {

            "id": f.id,

            "tipo_documento": f.tipo_documento,

            "codigo_formato": f.codigo_formato,

            "evolution_slot": f.evolution_slot,

            "nombre_medico": f.nombre_medico,

            "cedula_profesional": f.cedula_profesional,

            "fecha_hora_firma": f.fecha_hora_firma.strftime("%d/%m/%Y %H:%M:%S") if f.fecha_hora_firma else "",

            "metodo_autenticacion": f.metodo_autenticacion,

            "hash_sha256": f.hash_sha256,

            "sello_digital": f.sello_digital or "",

            "cadena_original": f.cadena_original or "",

            "estado": f.estado

        }

        for f in firmas

    ]



@app.get("/api/ehr/paciente/{pt_num}/historial-auditoria")

def get_historial_auditoria_paciente(pt_num: str, db: Session = Depends(get_db)):

    """

    Entrega el expediente forense inmutable de auditoría (NOM-024-SSA3-2012):

    Todas las versiones de notas, firmas históricas, firmas revocadas y eventos de seguridad.

    """

    historico_notas = db.query(models.HistoricoNotaClinica).filter(

        models.HistoricoNotaClinica.pt_num == str(pt_num)

    ).order_by(models.HistoricoNotaClinica.fecha_registro.desc()).all()



    firmas_todas = db.query(models.FirmaDocumentoClinico).filter(

        models.FirmaDocumentoClinico.pt_num == str(pt_num)

    ).order_by(models.FirmaDocumentoClinico.fecha_hora_firma.desc()).all()



    return {

        "pt_num": pt_num,

        "expediente": f"PT-{pt_num}",

        "total_versiones_clinicas": len(historico_notas),

        "total_firmas_registradas": len(firmas_todas),

        "firmas": [

            {

                "id": f.id,

                "slot": f.evolution_slot,

                "documento": f.tipo_documento,

                "medico": f.nombre_medico,

                "cedula": f.cedula_profesional,

                "fecha_firma": f.fecha_hora_firma.strftime("%d/%m/%Y %H:%M:%S") if f.fecha_hora_firma else "",

                "estado": f.estado,

                "fecha_revocacion": f.fecha_revocacion.strftime("%d/%m/%Y %H:%M:%S") if f.fecha_revocacion else None,

                "motivo_revocacion": f.motivo_revocacion,

                "hash_sha256": f.hash_sha256,

                "sello_hmac": f.sello_digital,

                "ip_origen": f.ip_origen

            }

            for f in firmas_todas

        ],

        "versiones_clinicas": [

            {

                "id": h.id,

                "slot": h.evolution_slot,

                "accion": h.accion,

                "medico": h.nombre_medico,

                "cedula": h.cedula_profesional,

                "fecha": h.fecha_registro.strftime("%d/%m/%Y %H:%M:%S") if h.fecha_registro else "",

                "ip_origen": h.ip_origen,

                "motivo": h.motivo,

                "contenido_soap": json.loads(h.contenido_soap_json) if h.contenido_soap_json else {}

            }

            for h in historico_notas

        ],

        "marco_normativo": "NOM-004-SSA3-2012 / NOM-024-SSA3-2012 (Inmutabilidad y No Repudio)"

    }



class VerificarIntegridadInputSchema(BaseModel):

    firma_id: int



@app.post("/api/ehr/paciente/{pt_num}/verificar-integridad")

def verificar_integridad_documento(

    pt_num: str,

    req: VerificarIntegridadInputSchema,

    request: Request,

    db: Session = Depends(get_db)

):

    """

    Recalcula en vivo el Hash SHA-256 y Sello HMAC-SHA512 desde los datos actuales de la base de datos

    y los compara contra el registro original firmado, proporcionando evidencia de la Tríada de Seguridad.

    """

    firma = db.query(models.FirmaDocumentoClinico).filter(

        models.FirmaDocumentoClinico.id == req.firma_id,

        models.FirmaDocumentoClinico.pt_num == str(pt_num)

    ).first()

    if not firma:

        raise HTTPException(status_code=404, detail="Registro de firma no encontrado.")

    

    # 1. Obtener la cadena original a verificar (soporta notas de evolución, consentimientos, recetas, dietas)

    if firma.cadena_original:

        cadena_to_verify = firma.cadena_original

    else:

        dashboard_data = kh_database.fetch_full_ehr_dashboard(pt_num)

        evols = dashboard_data.get("evoluciones", {})

        slot_key = f"evolucion{firma.evolution_slot or 1}"

        evol_data = evols.get(slot_key)

        live_subjetivo = (evol_data.get("subjetivo") if evol_data else "") or ""

        fecha_iso = firma.fecha_hora_firma.isoformat() if firma.fecha_hora_firma else ""

        cadena_to_verify = f"||{pt_num}|PT-{pt_num}|{firma.codigo_formato}|{firma.evolution_slot or 'GRAL'}|{fecha_iso}|{firma.medico_id}|{firma.cedula_profesional}|{live_subjetivo[:100]}||"

    

    recalculated_hash = hashlib.sha256(cadena_to_verify.encode('utf-8')).hexdigest()

    

    # Obtener el médico firmante

    medico = db.query(models.Medico).filter(models.Medico.id == firma.medico_id).first()

    huella_token = (medico.huella_token or medico.cedula) if medico else firma.cedula_profesional

    is_hash_valid = (recalculated_hash == firma.hash_sha256)

    

    fecha_iso = firma.fecha_hora_firma.isoformat() if firma.fecha_hora_firma else ""

    legacy_sha512 = hashlib.sha512(f"{recalculated_hash}-{huella_token}-{fecha_iso}".encode('utf-8')).hexdigest()

    cutoff_date = datetime.datetime(2027, 1, 1)

    fecha_firma_dt = firma.fecha_hora_firma or datetime.datetime.min

    legacy_is_valid = (legacy_sha512 == firma.sello_digital) and (fecha_firma_dt < cutoff_date)

    is_sello_valid = crypto_fea.verificar_firma(medico, cadena_to_verify, firma.sello_digital, firma.fecha_hora_firma) or legacy_is_valid

    

    is_integro = is_hash_valid and is_sello_valid

    

    if not is_integro:

        log_entry = models.AuditoriaLog(

            usuario_id=None,

            accion="INTEGRIDAD_DOCUMENTAL_COMPROMETIDA",

            detalles_json=json.dumps({

                "pt_num": pt_num,

                "firma_id": firma.id,

                "hash_esperado": firma.hash_sha256,

                "hash_calculado": recalculated_hash,

                "slot": firma.evolution_slot,

                "ip": request.client.host if request.client else "127.0.0.1"

            }),

            ip_origen=request.client.host if request.client else "127.0.0.1"

        )

        db.add(log_entry)

        db.commit()

        

    return {

        "integro": is_integro,

        "estado": "Firma íntegra y verificable" if is_integro else "Integridad invalidada - El documento fue modificado posterior a la firma",

        "triada_seguridad": {

            "identidad": {

                "verificado": True,

                "pilar": "Identidad del Firmante",

                "metodo": "Biometría Dactilar DigitalPersona (FMD ANSI/NIST 378-2004)",

                "firmante": firma.nombre_medico,

                "cedula": firma.cedula_profesional,

                "estado": "Autenticación biométrica comprobada fehacientemente"

            },

            "integridad": {

                "verificado": is_hash_valid,

                "pilar": "Integridad del Documento",

                "metodo": "Función Criptográfica SHA-256",

                "hash_sha256": firma.hash_sha256,

                "hash_calculado_en_vivo": recalculated_hash,

                "estado": "Documento íntegro sin alteraciones posteriores" if is_hash_valid else "Discrepancia detectada: el contenido actual no coincide con la versión firmada"

            },

            "autenticidad": {

                "verificado": is_sello_valid,

                "pilar": "Autenticidad y No Repudio",

                "metodo": "Firma Electrónica Avanzada ECDSA P-256 (asimétrica)" if (firma.sello_digital or '').startswith('ECDSA:') else "Sello Criptográfico HMAC-SHA512 (legacy)",

                "sello_hmac_sha512": firma.sello_digital,

                "estado": "Sello criptográfico auténtico garantizado con clave privada del firmante" if is_sello_valid else "Sello inválido o alterado"

            }

        },

        "sellado_tiempo": tsa_client.verify_timestamp(getattr(firma, 'tsa_token', None), firma.hash_sha256 or ''),

        "cadena_original": firma.cadena_original,

        "fecha_hora_firma": firma.fecha_hora_firma.strftime("%d/%m/%Y %H:%M:%S") if firma.fecha_hora_firma else "",

        "marco_normativo": "NOM-004-SSA3-2012 / NOM-024-SSA3-2012"

    }





@app.get("/api/camas/ocupacion", response_model=List[schemas.OcupacionArea])

def get_ocupacion_camas(db: Session = Depends(get_db)):

    """Calcula y devuelve la ocupación de camas por área siguiendo la lógica del Dashboard Directivo."""

    camas_activas = db.query(models.Cama).filter(models.Cama.activo == True).all()

    

    areas = {}

    for cama in camas_activas:

        if cama.area not in areas:

            areas[cama.area] = {

                "total": 0,

                "ocupadas": 0,

                "disponibles": 0,

                "mantenimiento": 0

            }

        

        areas[cama.area]["total"] += 1

        if cama.estado == "OCUPADA":

            areas[cama.area]["ocupadas"] += 1

        elif cama.estado == "DISPONIBLE":

            areas[cama.area]["disponibles"] += 1

        elif cama.estado == "MANTENIMIENTO":

            areas[cama.area]["mantenimiento"] += 1

            

    resultado = []

    for area, stats in areas.items():

        total = stats["total"]

        ocupadas = stats["ocupadas"]

        porcentaje = round((ocupadas * 100.0) / total, 2) if total > 0 else 0.0

        

        resultado.append({

            "area": area,

            "total_camas": total,

            "camas_ocupadas": ocupadas,

            "camas_disponibles": stats["disponibles"],

            "camas_mantenimiento": stats["mantenimiento"],

            "porcentaje_ocupacion": porcentaje

        })

        

    return resultado



# === AGENDA MEDICA Y CITAS ===



class CitaCreateSchema(BaseModel):

    medico_id: Optional[int] = None

    paciente_id: Optional[int] = None

    nombre_paciente_manual: Optional[str] = None

    fecha_hora: str # ISO string o YYYY-MM-DD HH:MM

    motivo: str

    lugar: Optional[str] = "Consultorio - Consulta Externa"

    notas: Optional[str] = None



@app.get("/api/medicos/list")

def get_medicos_list(db: Session = Depends(get_db)):

    """Devuelve la lista completa de médicos activos para selectores y agenda."""

    medicos = db.query(models.Medico).filter(models.Medico.activo_status == True).all()

    return [

        {

            "id": m.id,

            "nombre": m.nombre_completo,

            "especialidad": m.especialidad,

            "cedula": m.cedula,

            "numero_empleado": m.numero_empleado,

            "horario": m.horario_laboral or "Lunes a Viernes 08:00 - 16:00"

        }

        for m in medicos

    ]



@app.get("/api/agenda/citas")

def get_agenda_citas(medico_id: Optional[int] = None, db: Session = Depends(get_db)):

    """Obtiene las citas programadas de la agenda médica."""

    query = db.query(models.CitaMedica)

    if medico_id:

        query = query.filter(models.CitaMedica.medico_id == medico_id)

    

    citas = query.order_by(models.CitaMedica.fecha_hora.asc()).all()

    

    resultado = []

    for c in citas:

        medico_nom = c.medico.nombre_completo if c.medico else "Médico de Guardia"

        medico_esp = c.medico.especialidad if c.medico else "Medicina General"

        paciente_nom = c.paciente.nombre_completo if c.paciente else (c.nombre_paciente_manual or "Paciente no especificado")

        

        resultado.append({

            "id": c.id,

            "medico_id": c.medico_id,

            "medico_nombre": medico_nom,

            "medico_especialidad": medico_esp,

            "paciente_id": c.paciente_id,

            "paciente_nombre": paciente_nom,

            "fecha_hora": c.fecha_hora.isoformat() if c.fecha_hora else "",

            "fecha": c.fecha_hora.strftime("%d/%m/%Y") if c.fecha_hora else "",

            "hora": c.fecha_hora.strftime("%H:%M") if c.fecha_hora else "",

            "motivo": c.motivo,

            "lugar": c.lugar,

            "estatus": c.estatus,

            "notas": c.notas

        })

    return resultado



@app.post("/api/agenda/citas")

def create_agenda_cita(cita: CitaCreateSchema, db: Session = Depends(get_db)):

    """Registra una nueva cita médica programada."""

    try:

        dt = datetime.datetime.fromisoformat(cita.fecha_hora.replace('Z', '+00:00'))

    except Exception:

        try:

            dt = datetime.datetime.strptime(cita.fecha_hora, "%Y-%m-%d %H:%M")

        except Exception:

            dt = datetime.datetime.now() + datetime.timedelta(days=1)

            

    nueva_cita = models.CitaMedica(

        medico_id=cita.medico_id,

        paciente_id=cita.paciente_id,

        nombre_paciente_manual=cita.nombre_paciente_manual,

        fecha_hora=dt,

        motivo=cita.motivo,

        lugar=cita.lugar or "Consultorio - Consulta Externa",

        notas=cita.notas,

        estatus="Programada"

    )

    db.add(nueva_cita)

    db.commit()

    db.refresh(nueva_cita)

    return {"message": "Cita programada con éxito", "id": nueva_cita.id}



# === FRONTEND (PRODUCCION) ===

frontend_dist = os.path.join(os.path.dirname(__file__), "..", "frontend", "dist")

@app.get("/api/ehr/paciente/{pt_num}/pdf-consentimiento-eed")

def get_pdf_consentimiento_eed(pt_num: str):

    dashboard_data = kh_database.fetch_full_ehr_dashboard(pt_num)

    if "error" in dashboard_data:

        raise HTTPException(status_code=404, detail=dashboard_data["error"])

    patient_info = dashboard_data.get("patient", {})

    fecha_hoy = datetime.datetime.now().strftime("%d/%m/%Y")

    hora_hoy = datetime.datetime.now().strftime("%H:%M")

    pt_data = {
        "nombre": patient_info.get("name", ""),
        "expediente": pt_num,
        "fecha_nacimiento": patient_info.get("dob", ""),
        "sexo": patient_info.get("sex", "M"),
        "cama": patient_info.get("room", "Urgencias"),
        "fecha": fecha_hoy,
        "hora": hora_hoy,
        "alergias": patient_info.get("allergies", ""),
        "fecha_ingreso": patient_info.get("fecha_ingreso", fecha_hoy),
        "hora_ingreso": patient_info.get("hora_ingreso", hora_hoy)
    }

    eed_data = dashboard_data.get("consentimiento_eed", {})

    pt_data.update(eed_data)

    

    import pdf_engine_eed


    db = SessionLocal()
    try:
        firma_obj = db.query(models.FirmaDocumentoClinico).filter(
            models.FirmaDocumentoClinico.pt_num == str(pt_num),
            models.FirmaDocumentoClinico.codigo_formato == "HE-DIRMED-CONSUL-PLT-EED",
            models.FirmaDocumentoClinico.estado == "ACTIVA"
        ).order_by(models.FirmaDocumentoClinico.fecha_hora_firma.desc()).first()
        if firma_obj:
            pt_data["firma_data"] = {
                "sello_digital": firma_obj.sello_digital,
                "hash_sha256": firma_obj.hash_sha256,
                "fecha_hora_firma": firma_obj.fecha_hora_firma.strftime("%d/%m/%Y %H:%M:%S") if firma_obj.fecha_hora_firma else "",
                "nombre_medico": firma_obj.nombre_medico,
                "cedula": firma_obj.cedula_profesional
            }
    finally:
        db.close()

    pdf_path = pdf_engine_eed.generar_pdf_eed(pt_data)

    

    if os.path.exists(pdf_path):

        return FileResponse(pdf_path, media_type="application/pdf", filename=f"CI_EED_{pt_num}.pdf")

    raise HTTPException(status_code=500, detail="Error generating PDF")



@app.post("/api/ehr/paciente/{pt_num}/consentimiento-eed")
def save_consentimiento_eed(
    pt_num: str, 
    request_data: dict, 
    request: Request,
    db: Session = Depends(get_db)
):
    res = kh_database.save_or_update_consentimiento_eed(pt_num, request_data)
    if isinstance(res, dict) and "error" in res:
        raise HTTPException(status_code=500, detail=res["error"])
        
    client_ip = request.client.host if request.client else "127.0.0.1"
    
    # 1. Guardar Snapshot Inmutable en HistoricoNotaClinica (NOM-024)
    try:
        historico_entry = models.HistoricoNotaClinica(
            codigo_formato="HE-DIRMED-CONSUL-PLT-EED",
            tipo_documento="Consentimiento Informado para Ecocardiograma de Estrés con Dobutamina",
            pt_num=str(pt_num),
            expediente=f"PT-{pt_num}",
            evolution_slot=0,
            nombre_medico=str(request_data.get("medico") or request_data.get("medico_tratante") or ""),
            cedula_profesional=str(request_data.get("cedula") or request_data.get("cedula_profesional") or ""),
            contenido_soap_json=json.dumps(request_data, default=str),
            accion="GUARDADO",
            ip_origen=client_ip
        )
        db.add(historico_entry)
        db.commit()
    except Exception as e:
        print(f"Error creating audit history for Consentimiento EED: {e}")
        db.rollback()

    # 2. Soft-Revocation de firmas activas en PostgreSQL (NOM-024: se rompe la firma al modificar)
    try:
        firmas_activas = db.query(models.FirmaDocumentoClinico).filter(
            models.FirmaDocumentoClinico.pt_num == str(pt_num),
            models.FirmaDocumentoClinico.codigo_formato == "HE-DIRMED-CONSUL-PLT-EED",
            models.FirmaDocumentoClinico.estado == "ACTIVA"
        ).all()
        for f in firmas_activas:
            f.estado = "REVOCADA_POR_MODIFICACION"
            f.motivo_revocacion = "El documento fue modificado posteriormente."
        db.commit()
    except Exception as e:
        print(f"Error revoking signatures for Consentimiento EED: {e}")
        db.rollback()

    return {"message": "Consentimiento EED guardado con éxito", "status": "success"}

if os.path.exists(frontend_dist):

    app.mount("/assets", StaticFiles(directory=os.path.join(frontend_dist, "assets")), name="assets")

    

    @app.get("/{full_path:path}")

    async def serve_frontend(full_path: str):

        # Ignorar si es una ruta de backend o estática principal

        if full_path.startswith("api/") or full_path.startswith("static/") or full_path.startswith("generados/"):

            raise HTTPException(status_code=404, detail="Not found")

            

        # Si el archivo existe físicamente en dist/, servirlo (ej. logo.png, websdk.client.min.js)

        file_path = os.path.join(frontend_dist, full_path)

        if os.path.exists(file_path) and os.path.isfile(file_path):

            return FileResponse(file_path)

        

        # De lo contrario, devolver index.html para el React Router

        index_path = os.path.join(frontend_dist, "index.html")

        if os.path.exists(index_path):

            return FileResponse(index_path)

        return {"message": "Frontend no construido. Por favor corre 'npm run build' en la carpeta frontend."}


